"""
VoiceLibri Catalog Description Generator
=========================================
Fills empty/incomplete Description cells in voicelibri_assets_catalog.xlsx
using Gemini 2.5 Flash via Vertex AI REST API.

MODE 1: Complete truncated descriptions (add 1-2 words to finish sentence)
MODE 2: Generate missing descriptions from Keywords, FileName, Category, SubCategory

Features:
- Batch processing (50 rows per LLM call)
- Quality checks every 10 batches
- Style examples from existing complete descriptions
- Saves progress every 10 batches (crash-safe)
- Final validation pass
"""

import os
import sys
import json
import time
import re
import copy
import requests
import openpyxl
from google.oauth2 import service_account
from google.auth.transport.requests import Request as GoogleAuthRequest

# ── CONFIG ──────────────────────────────────────────────────────
CATALOG_PATH = os.path.join('soundscape', 'assets', 'voicelibri_assets_catalog.xlsx')
OUTPUT_PATH  = os.path.join('soundscape', 'assets', 'voicelibri_assets_catalog_described.xlsx')
SA_KEY_PATH  = os.path.join('apps', 'backend', '.gcsakey.json')
PROJECT_ID   = 'calmbridge-2'
LOCATION     = 'us-central1'
MODEL        = 'gemini-2.5-flash'
BATCH_SIZE   = 20
CHECK_EVERY  = 10  # validate quality every N batches
SAVE_EVERY   = 10  # save progress every N batches
SHOW_EVERY   = 5   # print sample Keywords vs Description every N batches
PROGRESS_FILE = os.path.join('scripts', '_desc_progress.json')

# Column indices (0-based)
COL_FILEID      = 0
COL_FILENAME    = 1
COL_DESCRIPTION = 2
COL_KEYWORDS    = 3
COL_CATEGORY    = 6
COL_SUBCATEGORY = 7

# ── AUTH ────────────────────────────────────────────────────────
def get_access_token():
    """Get OAuth2 access token from service account key."""
    creds = service_account.Credentials.from_service_account_file(
        SA_KEY_PATH,
        scopes=['https://www.googleapis.com/auth/cloud-platform']
    )
    creds.refresh(GoogleAuthRequest())
    return creds.token

def call_gemini(prompt: str, temperature: float = 0.7) -> str:
    """Call Gemini 2.5 Flash via Vertex AI REST API."""
    token = get_access_token()
    url = (
        f"https://{LOCATION}-aiplatform.googleapis.com/v1/"
        f"projects/{PROJECT_ID}/locations/{LOCATION}/"
        f"publishers/google/models/{MODEL}:generateContent"
    )
    body = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": temperature,
            "maxOutputTokens": 8192,
            "responseMimeType": "application/json",
            "thinkingConfig": {
                "thinkingBudget": 0
            },
        },
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    resp = requests.post(url, json=body, headers=headers, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    text = data["candidates"][0]["content"]["parts"][0]["text"]
    return text

# ── CLASSIFY ROWS ───────────────────────────────────────────────
def classify_description(desc):
    """Returns 'empty', 'partial', or 'complete'."""
    if desc is None or str(desc).strip() == '':
        return 'empty'
    s = str(desc).strip()
    # Truncated: ends without punctuation, or ends mid-word
    if s[-1] not in '.!?"\'':
        return 'partial'
    if len(s) < 15:
        return 'partial'
    return 'complete'

# ── COLLECT STYLE EXAMPLES ─────────────────────────────────────
def collect_style_examples(ws, headers, max_examples=30):
    """Collect diverse complete descriptions to teach the LLM the style."""
    examples = []
    seen_cats = set()
    seen_subs = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = [c.value for c in row]
        desc = vals[COL_DESCRIPTION]
        cat = str(vals[COL_CATEGORY] or '')
        sub = str(vals[COL_SUBCATEGORY] or '')

        if classify_description(desc) != 'complete':
            continue
        if len(str(desc).strip()) < 40:
            continue

        # Diversify: pick from different categories/subcategories
        cat_sub_key = f"{cat}|{sub}"
        if cat not in seen_cats or sub not in seen_subs:
            examples.append({
                'FileID': str(vals[COL_FILEID] or ''),
                'Filename': str(vals[COL_FILENAME] or ''),
                'Keywords': str(vals[COL_KEYWORDS] or ''),
                'Category': cat,
                'SubCategory': sub,
                'Description': str(desc).strip(),
            })
            seen_cats.add(cat)
            seen_subs.add(sub)

        if len(examples) >= max_examples:
            break

    return examples

# ── BUILD PROMPT ────────────────────────────────────────────────
def build_prompt(style_examples, batch_rows, mode_info="mixed"):
    """Build the LLM prompt with style examples and batch of rows to process."""

    # Use json.dumps for safe serialization (handles quotes, newlines, etc.)
    examples_text = json.dumps([
        {k: v for k, v in ex.items()}
        for ex in style_examples[:15]  # limit to 15 examples to keep prompt smaller
    ], indent=2, ensure_ascii=False)

    rows_data = [
        {'FileID': r['FileID'], 'Filename': r['Filename'],
         'Keywords': r['Keywords'], 'Category': r['Category'],
         'SubCategory': r['SubCategory'],
         'Description': r['Description'], 'Mode': r['Mode']}
        for r in batch_rows
    ]
    rows_text = json.dumps(rows_data, indent=2, ensure_ascii=False)

    prompt = f"""You are a literary author writing evocative scene descriptions — as if excerpted from a novel or screenplay. You paint pictures with words, describing what a listener would EXPERIENCE in that moment and place. You NEVER mention recordings, sound effects, microphones, studios, or audio files.

STUDY THESE EXAMPLES to understand the relationship between Keywords, Category, and how a description should feel:

{examples_text}

NOW PROCESS THESE ROWS. Each row has a "Mode" field:
- Mode "complete": The Description is TRUNCATED mid-sentence or mid-word. Finish it naturally. Add ONLY the minimum words needed (even just 1 word) to complete the sentence. Do NOT rephrase, rewrite, or extend beyond the original intent.
- Mode "generate": The Description is EMPTY. Write ONE vivid, literary sentence that paints the scene implied by the Keywords, as if describing a moment in a novel. Focus on atmosphere, place, and what one would hear and feel — NOT on technical audio terminology.

ROWS TO PROCESS:
{rows_text}

RULES:
1. Output ONLY a JSON array of objects with "FileID" and "Description" fields.
2. Every Description must be exactly ONE complete sentence ending with a period.
3. For "complete" mode: preserve the original text exactly, only append the missing ending.
4. FORBIDDEN words/phrases: "sound effect", "recording", "recorded", "ambience recorded", "captured", "audio", "sample", "field recording", "microphone", "studio". Write as a literary scene description instead.
5. For "generate" mode: write as if narrating a moment in a novel — vivid and atmospheric, never technical.
6. Use ENGLISH only.
7. STRICT GROUNDING — describe ONLY what the Keywords, Filename, Category, and SubCategory explicitly contain. Do NOT fabricate, infer, or add any details, imagery, or scene elements not present in the source data. Shorter descriptions are perfectly acceptable when the data is sparse.
7b. NEVER include the Category or SubCategory name (genre) in the description. The description must stand on its own without referencing its classification.
8. USE ALL MEANINGFUL KEYWORDS — every significant keyword from the Keywords field must appear in the description using the EXACT original word. Never substitute synonyms for keywords.
9. Output valid JSON array, nothing else.

OUTPUT FORMAT (JSON array only):
[
  {{"FileID": "00024", "Description": "The completed or generated description."}},
  ...
]"""
    return prompt

# ── QUALITY CHECK ───────────────────────────────────────────────
def quality_check(results, batch_rows):
    """Validate a batch of results. Returns (ok, issues).
    Only critical issues (empty, wrong start) count toward suspension.
    Multi-sentence warnings are logged but don't trigger suspension."""
    issues = []
    warnings = []

    if len(results) != len(batch_rows):
        issues.append(f"Row count mismatch: expected {len(batch_rows)}, got {len(results)}")
        return False, issues

    result_map = {str(r.get('FileID','')): r.get('Description','') for r in results}

    for row in batch_rows:
        fid = row['FileID']
        desc = result_map.get(fid, '')

        if not desc or desc.strip() == '':
            issues.append(f"FileID {fid}: empty description")
            continue

        s = desc.strip()
        # Must end with sentence-ending punctuation
        if s[-1] not in '.!?"\'':
            issues.append(f"FileID {fid}: doesn't end with punctuation: ...{s[-30:]}")

        # Multi-sentence: only a soft warning (many descriptions legitimately have periods)
        inner = s[:-1]
        period_count = inner.count('. ')
        if period_count > 2:
            warnings.append(f"FileID {fid}: possibly {period_count+1} sentences")

        # For partial completions, check we didn't replace original
        if row['Mode'] == 'complete' and row['Description']:
            orig_start = row['Description'][:50]
            if not s.startswith(orig_start[:30]):
                issues.append(f"FileID {fid}: partial completion changed original text")

    ok = len(issues) == 0
    all_notes = issues + [f"(warn) {w}" for w in warnings[:3]]
    return ok, all_notes

# ── PARSE LLM RESPONSE ─────────────────────────────────────────
def parse_llm_response(response_text):
    """Parse JSON array from LLM response, handling edge cases."""
    text = response_text.strip()
    # Remove markdown code blocks if present
    text = re.sub(r'^```\w*\n?', '', text)
    text = re.sub(r'\n?```\s*$', '', text)
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # Try to find JSON array in the text
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if match:
        candidate = match.group()
        # Fix trailing commas before ] (common LLM mistake)
        candidate = re.sub(r',\s*\]', ']', candidate)
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            pass
    # Last resort: try to fix truncated JSON by closing brackets
    if '[' in text and ']' not in text:
        # Find last complete object
        last_brace = text.rfind('}')
        if last_brace > 0:
            truncated = text[:last_brace+1] + ']'
            match2 = re.search(r'\[.*\]', truncated, re.DOTALL)
            if match2:
                candidate2 = re.sub(r',\s*\]', ']', match2.group())
                try:
                    return json.loads(candidate2)
                except json.JSONDecodeError:
                    pass
    raise json.JSONDecodeError('Could not extract JSON array from response', text, 0)

# ── SAVE / LOAD PROGRESS ───────────────────────────────────────
def save_progress(completed_ids):
    """Save set of completed FileIDs to disk."""
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(list(completed_ids), f)

def load_progress():
    """Load set of completed FileIDs from disk."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            return set(json.load(f))
    return set()

# ── MAIN ────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("VoiceLibri Catalog Description Generator")
    print("=" * 70)

    # Load workbook (full mode, not read_only, to allow editing)
    # Use the output file if it exists (it has accumulated descriptions from prior runs)
    load_path = OUTPUT_PATH if os.path.exists(OUTPUT_PATH) else CATALOG_PATH
    print(f"\n📂 Loading {load_path}...")
    wb = openpyxl.load_workbook(load_path)
    ws = wb.active
    total_rows = ws.max_row - 1  # exclude header
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"   Total rows: {total_rows}")
    print(f"   Headers: {headers}")

    # Load progress (for resume after crash)
    completed_ids = load_progress()
    if completed_ids:
        print(f"   ♻️ Resuming: {len(completed_ids)} rows already processed")

    # Collect style examples
    print("\n📚 Collecting style examples from complete descriptions...")
    # Need to re-read for examples (using a read-only copy for speed)
    wb_ro = openpyxl.load_workbook(load_path, read_only=True)
    ws_ro = wb_ro.active
    style_examples = collect_style_examples(ws_ro, headers, max_examples=30)
    wb_ro.close()
    print(f"   Collected {len(style_examples)} diverse style examples across categories")
    for ex in style_examples[:5]:
        print(f"   • [{ex['Category']}/{ex['SubCategory']}] {ex['Description'][:80]}...")

    # Identify rows to process
    print("\n🔍 Scanning for empty/partial descriptions...")
    rows_to_process = []
    for row_idx in range(2, ws.max_row + 1):
        fid = str(ws.cell(row_idx, COL_FILEID + 1).value or '')
        if fid in completed_ids:
            continue

        desc = ws.cell(row_idx, COL_DESCRIPTION + 1).value
        status = classify_description(desc)

        if status == 'complete':
            continue

        rows_to_process.append({
            'row_idx': row_idx,
            'FileID': fid,
            'Filename': str(ws.cell(row_idx, COL_FILENAME + 1).value or ''),
            'Keywords': str(ws.cell(row_idx, COL_KEYWORDS + 1).value or ''),
            'Category': str(ws.cell(row_idx, COL_CATEGORY + 1).value or ''),
            'SubCategory': str(ws.cell(row_idx, COL_SUBCATEGORY + 1).value or ''),
            'Description': str(desc or '').strip(),
            'Mode': 'complete' if status == 'partial' else 'generate',
        })

    empty_count = sum(1 for r in rows_to_process if r['Mode'] == 'generate')
    partial_count = sum(1 for r in rows_to_process if r['Mode'] == 'complete')
    print(f"   Empty: {empty_count}, Partial: {partial_count}, Total: {len(rows_to_process)}")

    if not rows_to_process:
        print("\n✅ All descriptions are complete! Nothing to do.")
        wb.close()
        return

    # Process in batches
    total_batches = (len(rows_to_process) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f"\n🚀 Processing {len(rows_to_process)} rows in {total_batches} batches of {BATCH_SIZE}")
    print(f"   Quality check every {CHECK_EVERY} batches")
    print(f"   Save progress every {SAVE_EVERY} batches")
    print()

    processed = 0
    failed_batches = 0
    batch_num = 0
    total_quality_issues = []

    for i in range(0, len(rows_to_process), BATCH_SIZE):
        batch = rows_to_process[i:i + BATCH_SIZE]
        batch_num += 1

        modes = f"({sum(1 for r in batch if r['Mode']=='generate')}gen, {sum(1 for r in batch if r['Mode']=='complete')}fix)"
        print(f"  📦 Batch {batch_num}/{total_batches} {modes}...", end=" ", flush=True)

        try:
            # Build and send prompt
            prompt = build_prompt(style_examples, batch)
            t0 = time.time()
            response = call_gemini(prompt, temperature=0.6)
            elapsed = time.time() - t0

            # Parse response
            results = parse_llm_response(response)
            print(f"✓ {len(results)} results ({elapsed:.1f}s)", end="")

            # Quality check on periodic intervals
            if batch_num % CHECK_EVERY == 0:
                ok, issues = quality_check(results, batch)
                if not ok:
                    print(f" ⚠️ QUALITY ISSUES:")
                    for issue in issues[:5]:
                        print(f"      {issue}")
                    total_quality_issues.extend(issues)
                    if len(issues) > len(batch) * 0.3:  # >30% failure rate
                        print(f"\n❌ SUSPENDED: Too many quality issues ({len(issues)}/{len(batch)})")
                        print(f"   Saving progress and stopping...")
                        save_progress(completed_ids)
                        wb.save(OUTPUT_PATH)
                        print(f"   Progress saved to {OUTPUT_PATH}")
                        return
                else:
                    print(f" ✅ quality OK", end="")

            print()

            # Build result map for this batch
            result_map = {str(r.get('FileID','')): r.get('Description','') for r in results}

            # Show sample Keywords → Description examples periodically
            if batch_num % SHOW_EVERY == 0:
                n_samples = 4 if batch_num % CHECK_EVERY == 0 else 2
                print(f"  ── Sample results (batch {batch_num}) ──")
                # Pick samples: prefer a mix of generate + complete modes
                samples_gen = [r for r in batch if r['Mode'] == 'generate'][:max(1, n_samples//2+1)]
                samples_fix = [r for r in batch if r['Mode'] == 'complete'][:max(1, n_samples//2)]
                samples = (samples_gen + samples_fix)[:n_samples]
                if not samples:
                    samples = batch[:n_samples]
                for s in samples:
                    fid = s['FileID']
                    new_desc = result_map.get(fid, '???')
                    kw = s['Keywords'][:80] + ('…' if len(s['Keywords']) > 80 else '')
                    mode_tag = '🆕' if s['Mode'] == 'generate' else '🔧'
                    print(f"    {mode_tag} [{s['Category']}/{s['SubCategory']}]")
                    print(f"       Keywords : {kw}")
                    if s['Mode'] == 'complete':
                        print(f"       Original : {s['Description'][:90]}{'…' if len(s['Description'])>90 else ''}")
                    print(f"       Result   : {new_desc[:120]}")
                print(f"  {'─' * 50}")

            # Write results back to worksheet
            for row_data in batch:
                fid = row_data['FileID']
                new_desc = result_map.get(fid, '')
                if new_desc and new_desc.strip():
                    ws.cell(row_data['row_idx'], COL_DESCRIPTION + 1).value = new_desc.strip()
                    completed_ids.add(fid)
                    processed += 1

            # Save progress periodically
            if batch_num % SAVE_EVERY == 0:
                print(f"  💾 Saving progress ({processed} processed so far)...")
                save_progress(completed_ids)
                wb.save(OUTPUT_PATH)

        except Exception as e:
            err_msg = str(e)[:200]
            print(f"❌ ERROR: {err_msg}")
            failed_batches += 1
            if failed_batches >= 5:
                print(f"\n❌ SUSPENDED: {failed_batches} consecutive failures")
                print(f"   Last error: {err_msg}")
                save_progress(completed_ids)
                wb.save(OUTPUT_PATH)
                print(f"   Progress saved to {OUTPUT_PATH}")
                return
            # Wait and retry with increasing delay
            delay = 10 * failed_batches
            print(f"   Waiting {delay}s before retry (attempt {failed_batches}/5)...")
            time.sleep(delay)
            continue

        # Reset failure counter on success
        failed_batches = 0

        # Rate limiting: small delay between batches
        time.sleep(1)

    # Final save
    print(f"\n💾 Saving final output to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    save_progress(completed_ids)

    # Final validation
    print(f"\n🔎 Final validation...")
    wb_check = openpyxl.load_workbook(OUTPUT_PATH, read_only=True)
    ws_check = wb_check.active
    still_empty = 0
    still_partial = 0
    for row in ws_check.iter_rows(min_row=2, max_row=ws_check.max_row):
        desc = row[COL_DESCRIPTION].value
        status = classify_description(desc)
        if status == 'empty': still_empty += 1
        elif status == 'partial': still_partial += 1
    wb_check.close()

    print(f"\n{'='*70}")
    print(f"✅ COMPLETE")
    print(f"   Processed: {processed} descriptions")
    print(f"   Still empty: {still_empty}")
    print(f"   Still partial: {still_partial}")
    print(f"   Quality issues found: {len(total_quality_issues)}")
    print(f"   Output: {OUTPUT_PATH}")
    print(f"{'='*70}")

    # Cleanup progress file on full completion
    if still_empty == 0 and still_partial == 0:
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)
            print("   🧹 Progress file cleaned up")

    wb.close()

if __name__ == '__main__':
    main()

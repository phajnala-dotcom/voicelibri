from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import urllib.parse

root = Path(r"C:\Users\hajna\VoiceLibri\soundscape\assets\pixabay")
root.mkdir(parents=True, exist_ok=True)

periods = ["prehistory","antiquity","middle ages","modernity","contemporary","timeless"]
music_genres = [
    "classical","folk","medieval","renaissance","baroque","chamber",
    "acoustic","cinematic","orchestral","ambient","jazz","blues",
    "world","latin","celtic"
]

moods = [
    "calm","dreamy","mysterious","tense","uplifting","melancholic",
    "solemn","heroic","playful","ominous","warm","majestic",
    "serene","wistful","energetic"
]

ambience_envs = [
    "nature","interior","city","farm","industrial","technology",
    "transport","weather","water","forest","mountain","coastal",
    "countryside","market","space"
]

ambience_sources = [
    "wind","rain","birds","traffic","crowd","machinery","water","river",
    "waves","fire","creaks","footsteps","echo","hum","thunder"
]

intensities = [
    "calm","soft","busy","stormy","distant","close","night","day",
    "low","medium","high","gentle","harsh","steady","swirling"
]

def build_music_search_url(query: str) -> str:
    return f"https://pixabay.com/music/search/{urllib.parse.quote(query)}/?order=ec"

def build_sound_search_url(query: str) -> str:
    return f"https://pixabay.com/sound-effects/search/{urllib.parse.quote(query)}/?order=ec"

wb = Workbook()
wb.remove(wb.active)

ws_music = wb.create_sheet("Music")
headers = [
    "ID","Section","Period","Genre","Mood1","Mood2","TargetFilename",
    "DurationSec","BitrateMinKbps","Stereo","DownloadsMin","LikesMin",
    "PixabayDirectURL","PixabayPageURL","Notes"
]
ws_music.append(headers)
for cell in ws_music[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

music_rows = []
idx = 0
# Weighted distribution: more recent periods have more files, timeless highest
music_period_targets = {
    "prehistory": 4,
    "antiquity": 8,
    "middle ages": 12,
    "modernity": 18,
    "contemporary": 26,
    "timeless": 32,
}

for p in periods:
    target = music_period_targets.get(p, 0)
    if target == 0:
        continue
    for i in range(target):
        g = music_genres[(idx + i) % len(music_genres)]
        mood1 = moods[(idx + i) % len(moods)]
        mood2 = moods[(idx + i + 3) % len(moods)]
        filename = f"{p}_{g}_{mood1}_{mood2}_{idx:03d}.mp3"
        music_rows.append([
            f"{idx:03d}","Music",p,g,mood1,mood2,filename,
            "","230","stereo","","","","",""
        ])
        idx += 1

# Ensure exactly 100 music rows
while len(music_rows) < 100:
    p = periods[-1]
    g = music_genres[idx % len(music_genres)]
    mood1 = moods[idx % len(moods)]
    mood2 = moods[(idx + 3) % len(moods)]
    filename = f"{p}_{g}_{mood1}_{mood2}_{idx:03d}.mp3"
    music_rows.append([
        f"{idx:03d}","Music",p,g,mood1,mood2,filename,
        "","230","stereo","","","","",""
    ])
    idx += 1

for row in music_rows:
    period, genre, mood1, mood2 = row[2], row[3], row[4], row[5]
    query = f"{genre} {mood1} {mood2} instrumental no vocals"
    page_url = build_music_search_url(query)
    row[12] = ""  # direct URL unavailable without audio API access
    row[13] = page_url
    row[14] = "EditorsChoice search; no vocals; verify acoustic/instrumental"
    ws_music.append(row)

ws_amb = wb.create_sheet("Ambience")
headers_amb = [
    "ID","Section","Environment","Source1","Source2","Mood/Intensity","TargetFilename",
    "DurationSec","BitrateMinKbps","Stereo","DownloadsMin","LikesMin",
    "PixabayDirectURL","PixabayPageURL","Notes"
]
ws_amb.append(headers_amb)
for cell in ws_amb[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

amb_rows = []
idx = 0
# Weighted distribution: emphasize nature, then interior, then city, then others
ambience_env_targets = {
    "nature": 120,
    "interior": 80,
    "city": 60,
    "farm": 30,
    "industrial": 25,
    "technology": 20,
    "transport": 20,
    "weather": 20,
    "water": 15,
    "forest": 10,
    "mountain": 8,
    "coastal": 5,
    "countryside": 4,
    "market": 2,
    "space": 1,
}

for env in ambience_envs:
    target = ambience_env_targets.get(env, 0)
    if target == 0:
        continue
    s_idx = 0
    while target > 0:
        s1 = ambience_sources[s_idx % len(ambience_sources)]
        s2 = ambience_sources[(s_idx + 3) % len(ambience_sources)]
        if s1 == s2:
            s_idx += 1
            continue
        mood = intensities[idx % len(intensities)]
        filename = f"{env}_{s1}_{s2}_{mood}_{idx:04d}.mp3"
        amb_rows.append([
            f"{idx:03d}","Ambience",env,s1,s2,mood,filename,
            "","230","stereo","","","","",""
        ])
        idx += 1
        s_idx += 1
        target -= 1

# Ensure exactly 400 ambience rows
while len(amb_rows) < 400:
    env = "nature"
    s1 = ambience_sources[idx % len(ambience_sources)]
    s2 = ambience_sources[(idx + 5) % len(ambience_sources)]
    if s1 == s2:
        s2 = ambience_sources[(idx + 7) % len(ambience_sources)]
    mood = intensities[idx % len(intensities)]
    filename = f"{env}_{s1}_{s2}_{mood}_{idx:04d}.mp3"
    amb_rows.append([
        f"{idx:03d}","Ambience",env,s1,s2,mood,filename,
        "","230","stereo","","","","",""
    ])
    idx += 1

for row in amb_rows:
    env, s1, s2, mood = row[2], row[3], row[4], row[5]
    query = f"{env} {s1} {s2} {mood} ambience"
    page_url = build_sound_search_url(query)
    row[12] = ""  # direct URL unavailable without audio API access
    row[13] = page_url
    row[14] = "EditorsChoice search; verify stereo 230kbps+"
    ws_amb.append(row)

for ws in (ws_music, ws_amb):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

out_path = root / "pixabay_download_plan.xlsx"
try:
    wb.save(out_path)
    print(out_path)
except PermissionError:
    alt_path = root / "pixabay_download_plan_v2.xlsx"
    wb.save(alt_path)
    print(alt_path)

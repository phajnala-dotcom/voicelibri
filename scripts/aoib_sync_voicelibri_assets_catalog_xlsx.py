"""
Sync voicelibri_assets_catalog.xlsx with aoib_ogg folder.

Finds OGG files on disk that are NOT in the catalog ('extras'),
reads their embedded metadata via ffprobe, and appends rows to the XLSX.

Usage:
    python scripts/aoib_sync_voicelibri_assets_catalog_xlsx.py              # dry-run audit
    python scripts/aoib_sync_voicelibri_assets_catalog_xlsx.py --apply      # append rows to a copy
    python scripts/aoib_sync_voicelibri_assets_catalog_xlsx.py --apply --in-place  # append in place (backup created)
"""

import argparse
import csv
import datetime as _dt
import json
import os
import shutil
import subprocess
import time
from dataclasses import dataclass
from pathlib import Path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(p: str) -> str:
    return p.replace('\\', '/').replace('//', '/')


def _strip_prefix(path: str, prefix: str) -> str:
    if path.lower().startswith(prefix.lower()):
        return path[len(prefix):]
    return path


def catalog_expected_rel_ogg(file_path: str) -> str:
    p = _norm(file_path).strip().strip('"')
    p = _strip_prefix(p, 'AIOB 4824/')
    if p.lower().endswith('.wav'):
        p = p[:-4] + '.ogg'
    return p


def file_path_to_rel_ogg(file_path: str) -> str:
    return catalog_expected_rel_ogg(file_path)


def _long_path(p: Path) -> str:
    s = str(p)
    if os.name == 'nt' and not s.startswith('\\\\?\\'):
        return '\\\\?\\' + s
    return s


def _strip_lp(s: str) -> str:
    return s[4:] if s.startswith('\\\\?\\') else s


def _parse_duration_to_seconds(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    if s.isdigit():
        return float(s)
    if ':' in s:
        parts = s.split(':')
        try:
            nums = [float(p) for p in parts]
        except ValueError:
            return None
        if len(nums) == 3:
            h, m, sec = nums
            return h * 3600 + m * 60 + sec
        if len(nums) == 2:
            m, sec = nums
            return m * 60 + sec
    return None


def _format_duration_mmss_ms(seconds: float | None) -> str:
    if seconds is None:
        return ''
    if seconds < 0:
        return ''
    total = int(seconds)
    ms = int(round((seconds - total) * 1000))
    if ms == 1000:
        total += 1
        ms = 0
    s = total % 60
    m = (total // 60) % 60
    h = total // 3600
    if h > 0:
        return f'{h:02d}:{m:02d}:{s:02d}.{ms:03d}'
    return f'{m:02d}:{s:02d}.{ms:03d}'


def _keywords_from_filename(filename: str) -> str:
    name = filename
    if name.lower().endswith('.wav'):
        name = name[:-4]
    name = name.replace('_', ' ').replace('-', ' ')
    parts = [p.strip() for p in name.split(',') if p.strip()]
    return ', '.join(parts)


def _clean_keywords(raw: str) -> str:
    if not raw:
        return ''
    replace_map = {
        'ambforst': 'forest ambience',
        'watrbubl': 'bubbling water',
        'watrflow': 'water flow',
        'watrdrip': 'dripping water',
    }
    drop_tokens = {'zapsplat', 'lr', 'slk'}
    out: list[str] = []
    seen: set[str] = set()
    for t in raw.replace(',', ' ').split():
        token = t.strip()
        low = token.lower()
        if not low:
            continue
        if low in drop_tokens or low.isdigit():
            continue
        if low in replace_map:
            for w in replace_map[low].split():
                if w not in seen:
                    out.append(w)
                    seen.add(w)
            continue
        # Normal token
        if low not in seen:
            out.append(token)
            seen.add(low)
    return ' '.join(out).strip()


def _infer_type_from_path(path_value: str) -> str:
    if path_value.startswith('realistic/'):
        return 'realistic'
    if path_value.startswith('cinematic_&_foley/'):
        return 'cinematic_&_foley'
    if path_value.startswith('music/'):
        return 'music'
    return ''


def _infer_location_from_filename(filename: str) -> str:
    return _infer_location_from_sources(filename, '')


def _subcategory_from_path_value(path_value: str) -> str:
    s = _norm(path_value).strip()
    for root in ('realistic/', 'cinematic_&_foley/', 'music/'):
        if s.startswith(root):
            rel = s[len(root):]
            parts = rel.split('/')
            if len(parts) > 1:
                candidate = parts[1]
                if candidate.lower().endswith(('.ogg', '.wav')):
                    return ''
                return candidate
            return ''
    return ''


def _normalize_location_key(value: str) -> str:
    return ''.join(ch.lower() for ch in value if ch.isalpha())


def _infer_location_from_sources(filename: str, path_value: str) -> str:
    geo_list = [
        'Australia', 'Austria', 'Belgium', 'Brazil', 'Cambodia', 'Canada', 'Chile', 'China', 'Croatia',
        'Czech Republic', 'Denmark', 'Egypt', 'Finland', 'France', 'Germany', 'Greece', 'Hungary',
        'Iceland', 'India', 'Indonesia', 'Ireland', 'Israel', 'Italy', 'Japan', 'Korea', 'South Korea',
        'North Korea', 'Latvia', 'Lithuania', 'Luxembourg', 'Malaysia', 'Mexico', 'Monaco',
        'Netherlands', 'New Zealand', 'Norway', 'Poland', 'Portugal', 'Romania', 'Russia', 'Saudi Arabia',
        'Singapore', 'Slovakia', 'Slovenia', 'South Africa', 'Spain', 'Sweden', 'Switzerland',
        'Taiwan', 'Thailand', 'Turkey', 'United Arab Emirates', 'United Kingdom', 'United States',
        'Vietnam',
        'Amsterdam', 'Athens', 'Auckland', 'Bangkok', 'Barcelona', 'Berlin', 'Bratislava', 'Brussels',
        'Budapest', 'Cairo', 'Copenhagen', 'Dublin', 'Edinburgh', 'Florence', 'Geneva', 'Helsinki',
        'Istanbul', 'Krakow', 'Lisbon', 'London', 'Los Angeles', 'Madrid', 'Melbourne', 'Milan',
        'Montreal', 'Moscow', 'Munich', 'Naples', 'New York', 'Oslo', 'Paris', 'Prague', 'Reykjavik',
        'Rome', 'San Francisco', 'Seoul', 'Shanghai', 'Singapore', 'Stockholm', 'Sydney', 'Tokyo',
        'Toronto', 'Venice', 'Vienna', 'Warsaw', 'Zurich'
    ]
    geo_map = {_normalize_location_key(v): v for v in geo_list}

    candidates: list[str] = []
    base = os.path.splitext(filename)[0]
    for part in base.split('_'):
        candidates.append(part.split('-')[0].split(',')[0].strip())
    if path_value:
        for seg in _norm(path_value).split('/'):
            candidates.append(seg.split('-')[0].split(',')[0].strip())

    joined = ' '.join(candidates)
    joined_key = _normalize_location_key(joined)
    for key, value in geo_map.items():
        if key and key in joined_key:
            return value

    for cand in candidates:
        if not cand:
            continue
        key = _normalize_location_key(cand)
        if key in geo_map:
            return geo_map[key]
    return ''


def ensure_catalog_csv(xlsx_path: Path, csv_path: Path) -> Path:
    if csv_path.exists() and csv_path.stat().st_mtime >= xlsx_path.stat().st_mtime:
        return csv_path

    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s:
            header = row_s
            header_row = r
            break

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open('w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            writer.writerow(row)

    wb.close()
    return csv_path


def _subcategory_from_relpath(rel_path: str) -> str:
    parts = rel_path.split('/')
    if len(parts) > 1:
        candidate = parts[1]
        if candidate.lower().endswith(('.ogg', '.wav')):
            return ''
        return candidate
    return ''


def _category_from_relpath(rel_path: str) -> str:
    parts = rel_path.split('/')
    if parts:
        return parts[0]
    return ''


def scan_asset_roots(roots: dict[str, Path]) -> dict[str, Path]:
    """Return map of root/rel -> full path for all ogg files under roots."""
    out: dict[str, Path] = {}
    for root_name, root_path in roots.items():
        root_lp = _long_path(root_path)
        root_str = _strip_lp(str(root_path))
        for dirpath, _dirs, files in os.walk(root_lp):
            for fname in files:
                if not fname.lower().endswith('.ogg'):
                    continue
                full = os.path.join(dirpath, fname)
                rel = _norm(os.path.relpath(_strip_lp(full), root_str))
                key = f'{root_name}/{rel}'
                out[key] = Path(_strip_lp(full))
    return out


def probe_duration_seconds(ogg_path: Path) -> float | None:
    try:
        cmd = [
            'ffprobe', '-v', 'quiet', '-print_format', 'json',
            '-show_format', '-show_streams', _long_path(ogg_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
        if proc.returncode != 0 or not proc.stdout.strip():
            return None
        data = json.loads(proc.stdout)
        streams = data.get('streams', [])
        if streams:
            s = streams[0]
            dur = s.get('duration')
            if dur:
                return float(dur)
        fmt = data.get('format', {})
        dur = fmt.get('duration')
        if dur:
            return float(dur)
    except Exception:
        return None
    return None


def _split_recmedium_microphone(raw: str) -> tuple[str, str]:
    if not raw:
        return '', ''
    parts = [p.strip() for p in raw.split(',') if p.strip()]
    if len(parts) >= 2:
        return parts[0], ', '.join(parts[1:])
    return '', raw.strip()


# ---------------------------------------------------------------------------
# Inventory
# ---------------------------------------------------------------------------

def read_csv_filepaths(csv_path: Path) -> set[str]:
    """Read CSV catalog and return set of expected OGG relpaths (fast)."""
    out: set[str] = set()
    with csv_path.open('r', encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            fp = (row.get('FilePath') or '').strip()
            if fp:
                out.add(catalog_expected_rel_ogg(fp))
    return out


def scan_ogg_relpaths(root: Path) -> list[str]:
    """Scan all OGG files under root, return sorted relpaths."""
    root_lp = _long_path(root)
    root_str = _strip_lp(str(root))
    seen: set[str] = set()
    out: list[str] = []
    for dirpath, _dirs, files in os.walk(root_lp):
        for fname in files:
            if not fname.lower().endswith('.ogg'):
                continue
            full = os.path.join(dirpath, fname)
            key = full.lower()
            if key in seen:
                continue
            seen.add(key)
            rel = _norm(os.path.relpath(_strip_lp(full), root_str))
            out.append(rel)
    out.sort(key=str.lower)
    return out


# ---------------------------------------------------------------------------
# Metadata extraction via ffprobe
# ---------------------------------------------------------------------------

@dataclass
class OggMeta:
    rel_path: str = ''
    filename_wav: str = ''
    file_path_wav: str = ''
    duration_s: float | None = None
    duration_str: str = ''
    category: str = ''
    subcategory: str = ''
    description: str = ''
    keywords: str = ''
    location: str = ''
    microphone: str = ''
    track_year: str = ''
    rec_medium: str = ''
    rec_type: str = ''
    channels: int = 0
    sample_rate: int = 0
    error: str = ''


def _fmt_duration(secs: float) -> str:
    m, s = divmod(int(secs), 60)
    h, m = divmod(m, 60)
    if h:
        return f'{h}:{m:02d}:{s:02d}'
    return f'{m}:{s:02d}'


def extract_metadata(ogg_path: Path, rel_path: str) -> OggMeta:
    """Extract as much metadata as possible from an OGG file."""
    meta = OggMeta(rel_path=rel_path)

    # Derive from path
    parts = rel_path.split('/')
    meta.category = parts[0] if parts else ''
    meta.subcategory = _subcategory_from_relpath(rel_path)
    basename = os.path.basename(rel_path)
    meta.filename_wav = (basename[:-4] + '.wav') if basename.lower().endswith('.ogg') else basename
    wav_rel = (rel_path[:-4] + '.wav') if rel_path.lower().endswith('.ogg') else rel_path
    meta.file_path_wav = f'AIOB 4824/{wav_rel}'

    # ffprobe for embedded metadata
    try:
        cmd = [
            'ffprobe', '-v', 'quiet', '-print_format', 'json',
            '-show_format', '-show_streams', _long_path(ogg_path),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
        if proc.returncode == 0 and proc.stdout.strip():
            data = json.loads(proc.stdout)

            # Stream info
            streams = data.get('streams', [])
            if streams:
                s = streams[0]
                meta.channels = int(s.get('channels', 0))
                meta.sample_rate = int(s.get('sample_rate', 0))
                dur = s.get('duration')
                if dur:
                    meta.duration_s = float(dur)
                    meta.duration_str = _fmt_duration(meta.duration_s)

                tags = s.get('tags', {})
                # Description from comment or ISBJ
                meta.description = (tags.get('comment') or tags.get('ISBJ') or '').strip()
                # Keywords from IKEY (used only as fallback)
                meta.keywords = (tags.get('IKEY') or '').strip()
                # Microphone / recorder from encoded_by or IARL
                raw_mic = (tags.get('encoded_by') or tags.get('IARL') or '').strip()
                meta.rec_medium, meta.microphone = _split_recmedium_microphone(raw_mic)
                # Year from TORY or date
                meta.track_year = (tags.get('TORY') or tags.get('date') or '').strip()
                # Genre → override category if available
                genre = (tags.get('genre') or '').strip()
                if genre:
                    meta.category = genre

            # Fallback duration from format
            if meta.duration_s is None:
                fmt = data.get('format', {})
                dur = fmt.get('duration')
                if dur:
                    meta.duration_s = float(dur)
                    meta.duration_str = _fmt_duration(meta.duration_s)

            # RecType is not used; keep channels/sample_rate only if needed later.

    except Exception as e:
        meta.error = str(e)

    return meta


# ---------------------------------------------------------------------------
# XLSX append
# ---------------------------------------------------------------------------

def _sanitize(val: object) -> object:
    """Strip illegal XML characters that openpyxl rejects."""
    if not isinstance(val, str):
        return val
    import re
    # Remove ASCII control chars except tab/newline/carriage-return
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', val)


def update_extras_in_xlsx(xlsx_path: Path, extras_rel: set[str]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Find header
    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s and 'Filename' in row_s:
            header = row_s
            header_row = r
            break

    def col(name: str) -> int | None:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    c_filename = col('Filename')
    c_keywords = col('Keywords')
    c_duration = col('Duration')
    c_subcategory = col('SubCategory')
    c_microphone = col('Microphone')
    c_recmedium = col('RecMedium')
    c_filepath = col('FilePath')

    updated = 0
    for r in range(header_row + 1, ws.max_row + 1):
        fp = ws.cell(row=r, column=c_filepath).value if c_filepath else None
        if not fp:
            continue
        rel = file_path_to_rel_ogg(str(fp))
        if rel not in extras_rel:
            continue

        filename = ws.cell(row=r, column=c_filename).value if c_filename else None
        if not filename:
            filename = os.path.basename(rel)[:-4] + '.wav'

        # Keywords from filename
        if c_keywords:
            ws.cell(row=r, column=c_keywords).value = _sanitize(
                _clean_keywords(_keywords_from_filename(str(filename)))
            )

        # Duration normalized
        if c_duration:
            dur_val = ws.cell(row=r, column=c_duration).value
            dur_sec = _parse_duration_to_seconds(dur_val)
            if dur_sec is not None:
                ws.cell(row=r, column=c_duration).value = _sanitize(_fmt_duration(dur_sec))

        # SubCategory from category subfolder (only if missing)
        if c_subcategory:
            current = ws.cell(row=r, column=c_subcategory).value
            if not current:
                ws.cell(row=r, column=c_subcategory).value = _sanitize(_subcategory_from_relpath(rel))

        # Split Microphone into RecMedium + Microphone (override old RecMedium like OGG Opus)
        if c_microphone:
            raw_mic = ws.cell(row=r, column=c_microphone).value or ''
            rec_medium, mic = _split_recmedium_microphone(str(raw_mic))
            if mic:
                ws.cell(row=r, column=c_microphone).value = _sanitize(mic)
            if c_recmedium:
                if rec_medium:
                    ws.cell(row=r, column=c_recmedium).value = _sanitize(rec_medium)
                else:
                    current = ws.cell(row=r, column=c_recmedium).value or ''
                    if 'ogg' in str(current).lower() or 'opus' in str(current).lower():
                        ws.cell(row=r, column=c_recmedium).value = ''

        updated += 1

    wb.save(xlsx_path)
    wb.close()
    print(f'  Updated {updated} existing extra rows')


def update_catalog_structure(
    xlsx_path: Path,
    roots: dict[str, Path],
) -> None:
    import openpyxl

    inventory = scan_asset_roots(roots)
    inventory_keys = set(inventory.keys())

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Find header
    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s and 'Filename' in row_s:
            header = row_s
            header_row = r
            break

    def col(name: str) -> int | None:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    c_fileid = col('FileID') or col('RecID')
    c_filename = col('Filename')
    c_description = col('Description')
    c_keywords = col('Keywords')
    c_duration = col('Duration')
    c_category = col('Category')
    c_subcategory = col('SubCategory')
    c_location = col('Location')
    c_microphone = col('Microphone')
    c_trackyear = col('TrackYear')
    c_recmedium = col('RecMedium')
    c_filepath = col('FilePath')
    c_type = col('Type')

    # Find max FileID
    next_id = 1
    if c_fileid:
        max_id = 0
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=c_fileid).value
            try:
                n = int(str(v).strip())
                if n > max_id:
                    max_id = n
            except (ValueError, TypeError):
                pass
        next_id = max_id + 1

    catalog_keys: set[str] = set()
    updated_paths = 0
    missing_files = 0

    for r in range(header_row + 1, ws.max_row + 1):
        fp = ws.cell(row=r, column=c_filepath).value if c_filepath else None
        if not fp:
            continue
        fp_s = _norm(str(fp).strip())

        # Map old AIOB paths to new roots
        if fp_s.startswith('AIOB 4824/'):
            rel = fp_s[len('AIOB 4824/'):]
            if rel.lower().endswith('.wav'):
                rel = rel[:-4] + '.ogg'
            found = None
            for root_name in roots.keys():
                key = f'{root_name}/{rel}'
                if key in inventory_keys:
                    found = key
                    break
            if found:
                ws.cell(row=r, column=c_filepath).value = _sanitize(found)
                catalog_keys.add(found)
                updated_paths += 1
                if c_type:
                    ws.cell(row=r, column=c_type).value = _sanitize(_infer_type_from_path(found))
                if c_keywords:
                    kw = ws.cell(row=r, column=c_keywords).value or ''
                    ws.cell(row=r, column=c_keywords).value = _sanitize(_clean_keywords(str(kw)))
                if c_subcategory:
                    current = ws.cell(row=r, column=c_subcategory).value
                    if not current:
                        ws.cell(row=r, column=c_subcategory).value = _sanitize(_subcategory_from_path_value(found))
                if c_location:
                    current = ws.cell(row=r, column=c_location).value
                    if not current and c_filename:
                        fname = ws.cell(row=r, column=c_filename).value or ''
                        loc = _infer_location_from_sources(str(fname), found)
                        if loc:
                            ws.cell(row=r, column=c_location).value = _sanitize(loc)
            else:
                missing_files += 1
        else:
            # Already new root path
            if fp_s in inventory_keys:
                catalog_keys.add(fp_s)
                if c_type:
                    ws.cell(row=r, column=c_type).value = _sanitize(_infer_type_from_path(fp_s))
                if c_keywords:
                    kw = ws.cell(row=r, column=c_keywords).value or ''
                    ws.cell(row=r, column=c_keywords).value = _sanitize(_clean_keywords(str(kw)))
                if c_subcategory:
                    current = ws.cell(row=r, column=c_subcategory).value
                    if not current:
                        ws.cell(row=r, column=c_subcategory).value = _sanitize(_subcategory_from_path_value(fp_s))
                if c_location:
                    current = ws.cell(row=r, column=c_location).value
                    if not current and c_filename:
                        fname = ws.cell(row=r, column=c_filename).value or ''
                        loc = _infer_location_from_sources(str(fname), fp_s)
                        if loc:
                            ws.cell(row=r, column=c_location).value = _sanitize(loc)
            else:
                missing_files += 1

    # Append new rows for files not in catalog
    new_keys = sorted(inventory_keys - catalog_keys)
    added = 0
    if new_keys:
        start_row = ws.max_row + 1
        for i, key in enumerate(new_keys):
            row = start_row + i
            root_name, rel = key.split('/', 1)
            full_path = inventory[key]
            base = os.path.basename(rel)
            filename_wav = base[:-4] + '.wav' if base.lower().endswith('.ogg') else base

            def _set(c: int | None, val: object) -> None:
                if c is not None and val not in (None, ''):
                    ws.cell(row=row, column=c).value = _sanitize(val)

            if c_fileid:
                ws.cell(row=row, column=c_fileid).value = next_id
                next_id += 1
            _set(c_filename, filename_wav)
            # Do NOT fill Description
            if c_description:
                ws.cell(row=row, column=c_description).value = ''
            _set(c_keywords, _clean_keywords(_keywords_from_filename(filename_wav)))
            dur = probe_duration_seconds(full_path)
            _set(c_duration, _format_duration_mmss_ms(dur))
            _set(c_category, _category_from_relpath(rel))
            _set(c_subcategory, _subcategory_from_relpath(rel))
            if c_location:
                loc = _infer_location_from_sources(filename_wav, key)
                ws.cell(row=row, column=c_location).value = _sanitize(loc) if loc else ''
            if c_microphone:
                ws.cell(row=row, column=c_microphone).value = ''
            if c_trackyear:
                ws.cell(row=row, column=c_trackyear).value = ''
            if c_recmedium:
                ws.cell(row=row, column=c_recmedium).value = ''
            if c_type:
                ws.cell(row=row, column=c_type).value = _sanitize(_infer_type_from_path(key))
            _set(c_filepath, key)
            added += 1

    wb.save(xlsx_path)
    wb.close()
    print(f'  Updated paths: {updated_paths}')
    print(f'  Missing files referenced: {missing_files}')
    print(f'  Added new rows: {added}')


def update_fields_from_paths(xlsx_path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s and 'Filename' in row_s:
            header = row_s
            header_row = r
            break

    def col(name: str) -> int | None:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    c_filename = col('Filename')
    c_keywords = col('Keywords')
    c_subcategory = col('SubCategory')
    c_location = col('Location')
    c_filepath = col('FilePath')

    updated = 0
    for r in range(header_row + 1, ws.max_row + 1):
        fp = ws.cell(row=r, column=c_filepath).value if c_filepath else None
        if not fp:
            continue
        fp_s = _norm(str(fp).strip())

        if c_keywords:
            kw = ws.cell(row=r, column=c_keywords).value or ''
            ws.cell(row=r, column=c_keywords).value = _sanitize(_clean_keywords(str(kw)))

        if c_subcategory:
            current = ws.cell(row=r, column=c_subcategory).value
            if not current:
                ws.cell(row=r, column=c_subcategory).value = _sanitize(_subcategory_from_path_value(fp_s))

        if c_location and c_filename:
            current = ws.cell(row=r, column=c_location).value
            if not current:
                fname = ws.cell(row=r, column=c_filename).value or ''
                loc = _infer_location_from_sources(str(fname), fp_s)
                if loc:
                    ws.cell(row=r, column=c_location).value = _sanitize(loc)

        updated += 1

    wb.save(xlsx_path)
    wb.close()
    print(f'  Updated fields for {updated} rows')


def append_to_xlsx(
    xlsx_path: Path,
    extras: list[OggMeta],
    in_place: bool,
    out_path: Path | None,
) -> Path:
    import openpyxl

    target = xlsx_path

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Find header
    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s and 'Filename' in row_s:
            header = row_s
            header_row = r
            break

    def col(name: str) -> int | None:
        try:
            return header.index(name) + 1
        except ValueError:
            return None

    c_fileid = col('FileID') or col('RecID')
    c_filename = col('Filename')
    c_description = col('Description')
    c_keywords = col('Keywords')
    c_duration = col('Duration')
    c_category = col('Category')
    c_subcategory = col('SubCategory')
    c_location = col('Location')
    c_microphone = col('Microphone')
    c_trackyear = col('TrackYear')
    c_recmedium = col('RecMedium')
    c_filepath = col('FilePath')

    # Find max FileID
    next_id = 1
    if c_fileid:
        max_id = 0
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=c_fileid).value
            try:
                n = int(str(v).strip())
                if n > max_id:
                    max_id = n
            except (ValueError, TypeError):
                pass
        next_id = max_id + 1

    # Append rows
    start_row = ws.max_row + 1
    for i, m in enumerate(extras):
        row = start_row + i

        def _set(c: int | None, val: object) -> None:
            if c and val:
                ws.cell(row=row, column=c).value = _sanitize(val)

        if c_fileid:
            ws.cell(row=row, column=c_fileid).value = next_id
            next_id += 1
        _set(c_filename, m.filename_wav)
        _set(c_description, m.description)
        _set(c_keywords, _keywords_from_filename(m.filename_wav))
        dur = m.duration_s if m.duration_s is not None else _parse_duration_to_seconds(m.duration_str)
        _set(c_duration, _fmt_duration(dur) if dur is not None else '')
        _set(c_category, m.category)
        _set(c_subcategory, m.subcategory)
        _set(c_location, m.location)
        _set(c_microphone, m.microphone)
        _set(c_trackyear, m.track_year)
        _set(c_recmedium, m.rec_medium)
        _set(c_filepath, m.file_path_wav)

    wb.save(target)
    wb.close()
    return target


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description='Sync voicelibri_assets_catalog.xlsx with aoib_ogg folder')
    ap.add_argument('--apply', action='store_true', help='Actually append rows to XLSX (default: dry-run audit only)')
    ap.add_argument('--in-place', action='store_true', help='Modify XLSX in place')
    ap.add_argument('--update-extras', action='store_true', help='Update existing extra rows (keywords/subcategory/duration/recmedium)')
    ap.add_argument('--update-structure', action='store_true', help='Update catalog for new asset roots and add new files')
    ap.add_argument('--update-fields', action='store_true', help='Improve keywords, fill missing subcategory and location')
    ap.add_argument('--use-csv-cache', action='store_true', help='Convert XLSX to CSV for faster reads (default behavior)')
    args = ap.parse_args()

    repo = Path(__file__).resolve().parents[1]
    csv_path = repo / 'soundscape' / 'assets' / 'voicelibri_assets_catalog.csv'
    xlsx_path = repo / 'soundscape' / 'assets' / 'voicelibri_assets_catalog.xlsx'
    aoib_ogg = repo / 'soundscape' / 'assets' / 'aoib_ogg'
    roots = {
        'realistic': repo / 'soundscape' / 'assets' / 'realistic',
        'cinematic': repo / 'soundscape' / 'assets' / 'cinematic',
        'music': repo / 'soundscape' / 'assets' / 'music',
        'SFX': repo / 'soundscape' / 'assets' / 'SFX',
    }

    extras: list[str] = []
    if not args.update_structure and not args.update_fields:
        print('Reading CSV catalog (fast)...', flush=True)
        csv_path = ensure_catalog_csv(xlsx_path, csv_path)
        t0 = time.time()
        expected = read_csv_filepaths(csv_path)
        print(f'  {len(expected)} expected paths  ({time.time() - t0:.1f}s)', flush=True)

        print('Scanning OGG files on disk...', flush=True)
        t0 = time.time()
        actual = scan_ogg_relpaths(aoib_ogg)
        actual_set = set(actual)
        print(f'  {len(actual)} OGG files  ({time.time() - t0:.1f}s)', flush=True)

        extras = sorted(actual_set - expected)
        print(f'  Extra (on disk, not in catalog): {len(extras)}', flush=True)

    if not extras and not args.update_extras and not args.update_structure and not args.update_fields:
        print('Nothing to add — catalog is already in sync.')
        return

    metas: list[OggMeta] = []
    if extras and args.apply:
        print(f'Extracting metadata from {len(extras)} extra files...', flush=True)
        t0 = time.time()
        for i, rel in enumerate(extras, 1):
            ogg_path = aoib_ogg / rel.replace('/', os.sep)
            m = extract_metadata(ogg_path, rel)
            metas.append(m)
            if i % 25 == 0 or i == len(extras):
                print(f'  {i}/{len(extras)} metadata extracted', flush=True)
        print(f'  Done ({time.time() - t0:.1f}s)', flush=True)

    # Print summary
    has_desc = sum(1 for m in metas if m.description)
    has_kw = sum(1 for m in metas if m.keywords)
    has_dur = sum(1 for m in metas if m.duration_str)
    has_mic = sum(1 for m in metas if m.microphone)
    has_year = sum(1 for m in metas if m.track_year)
    if metas:
        print(f'\nMetadata coverage ({len(metas)} files):')
        print(f'  Description: {has_desc}  Keywords: {has_kw}  Duration: {has_dur}  Microphone: {has_mic}  Year: {has_year}')

    if args.update_extras:
        update_extras_in_xlsx(xlsx_path, set(extras))

    if args.update_structure:
        update_catalog_structure(xlsx_path, roots)

    if args.update_fields:
        update_fields_from_paths(xlsx_path)

    if args.apply:
        print('\nAppending rows to XLSX...', flush=True)
        out = append_to_xlsx(xlsx_path, metas, in_place=args.in_place, out_path=None)
        print(f'  Wrote: {out}')
        print(f'  Added {len(metas)} rows')
    else:
        print(f'\nDry run — use --apply to append {len(metas)} rows to {xlsx_path.name}')
        print('Sample extras:')
        for m in metas[:5]:
            desc_preview = f'  desc={m.description[:60]}...' if m.description else ''
            print(f'  {m.rel_path}  dur={m.duration_str}  cat={m.category}{desc_preview}')


if __name__ == '__main__':
    main()

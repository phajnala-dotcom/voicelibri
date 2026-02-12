"""Soundscape catalog Phase 2 generator (rule-based generalization).

- Reads soundscape/assets/catalog.xlsx (sheet "Metadata, tags").
- Infers Type, Texture, Genre, Era, Location, Mood1, Mood2 where missing,
  using deterministic, whole-word, context-aware heuristics based on
  filenames, asset IDs and folder structure.
- Never overwrites non-empty cells (except for a few very conservative
  bug-fix patterns like removing "war" as a type/genre).
- Can optionally export updated Excel to JSON catalog and manifest.

This script is intentionally conservative: it prefers leaving a field empty
rather than guessing when context is ambiguous. It also enforces the key
rules:

- No word waste: map meaningful words into the MOST appropriate single field.
- No duplicates across semantic fields (Type / Genre / Texture / Moods).
- Horror as Genre (Type only as last-resort fallback).
- "war" is never used as Type/Genre; mapped to action/historical.
- City/location names treated as Location only; never inferred as Type.
- Whole-word matching only ("Newcastle" will NOT match "castle").
- "short"/"long"/"loop"/"stereo" etc. treated as technical metadata, not
  content categories.

Usage examples (from repository root):

  .venv/Scripts/python.exe scripts/generate_catalog_v2.py \
      --excel soundscape/assets/catalog.xlsx

Dry-run (only summary, no modifications):

  .venv/Scripts/python.exe scripts/generate_catalog_v2.py \
      --excel soundscape/assets/catalog.xlsx --dry-run

Export JSON + manifest AFTER you manually validate/lock the Excel:

  .venv/Scripts/python.exe scripts/generate_catalog_v2.py \
      --excel soundscape/assets/catalog.xlsx \
      --export-json --export-manifest

"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook

try:
    # Audio metadata reader (ID3/Vorbis/etc.)
    from mutagen import File as MutagenFile  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    MutagenFile = None


EXCEL_SHEET_NAME = "Metadata, tags"


def normalize_text(text: str) -> str:
    """Normalize text to a tokenization-friendly form.

    - Replace common separators with spaces.
    - Split camelCase / PascalCase boundaries.
    - Collapse multiple spaces.
    """

    # Ensure string
    text = str(text)

    # Insert spaces on camelCase boundaries
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)

    # Replace separators with spaces
    text = re.sub(r"[\-_,.;:/\\()\[\]]+", " ", text)

    # Collapse multiple spaces
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenize(text: str) -> List[str]:
    """Tokenize string into lowercased words with whole-word semantics.

    This is critical to avoid bugs like matching "castle" inside
    "Newcastle". We only ever match full tokens, not substrings.
    """

    if not text:
        return []
    norm = normalize_text(text)
    if not norm:
        return []
    return [t.lower() for t in norm.split(" ") if t]


def build_token_context(asset_id: str, filename: str) -> Tuple[List[str], Set[str]]:
    """Build ordered tokens list + a quick membership set from AssetID + Filename.

    AssetID path segments and filename (without extension) are both used.
    """

    base = filename.rsplit(".", 1)[0] if filename else ""
    combined = f"{asset_id} {base}" if asset_id else base
    tokens = tokenize(combined)
    return tokens, set(tokens)


def extract_metadata_tokens(base_dir: Path, asset_id: str, filename: str) -> Tuple[List[str], Set[str]]:
    """Extract tokens from embedded audio metadata (tags) using mutagen.

    Tags are *second-priority* compared to filenames/paths. We use them to:
    - Fill empty cells when filenames were insufficient.
    - Optionally *augment* existing cells by appending new, non-duplicate
      words (never overwriting filename-derived content).
    """

    if MutagenFile is None:
        return [], set()

    if not filename:
        return [], set()

    # Derive relative folder from AssetID (everything before last '/').
    rel = (asset_id or "").replace("\\", "/")
    folder = rel.rsplit("/", 1)[0] if "/" in rel else ""
    audio_path = base_dir / folder / filename

    if not audio_path.exists():
        return [], set()

    try:
        audio = MutagenFile(str(audio_path))
    except Exception:
        return [], set()

    if not audio or not getattr(audio, "tags", None):
        return [], set()

    parts: List[str] = []
    try:
        for value in audio.tags.values():  # type: ignore[union-attr]
            if isinstance(value, (list, tuple)):
                for v in value:
                    if isinstance(v, str):
                        parts.append(v)
            elif isinstance(value, str):
                parts.append(value)
    except Exception:
        # Be conservative; if tags are weird, just ignore them.
        return [], set()

    if not parts:
        return [], set()

    text = " ".join(parts)
    tokens = tokenize(text)
    return tokens, set(tokens)


# --- Lexicons --------------------------------------------------------------

# Technical / metadata terms that should NOT be used as semantic labels
TECHNICAL_WORDS: Set[str] = {
    "short",
    "long",
    "medium",
    "loop",
    "loops",
    "mono",
    "stereo",
    "version",
    "mix",
    "remix",
    "master",
    "take",
    "edit",
    "cut",
    "alt",
    "underscore",
    "stems",
    "intro",
    "outro",
    "ending",
    "end",
    "30",
    "60",
    "15",
    "30sec",
    "60sec",
    "sfx",
}

# Environment / setting cues for Type
NATURE_TYPES: Dict[str, str] = {
    "forest": "forest",
    "rainforest": "forest",
    "jungle": "jungle",
    "mountain": "mountain",
    "mountains": "mountain",
    "desert": "desert",
    "beach": "beach",
    "coast": "coast",
    "coastal": "coast",
    "ocean": "ocean",
    "sea": "sea",
    "river": "river",
    "stream": "river",
    "creek": "river",
    "waterfall": "waterfall",
    "lake": "lake",
    "cave": "cave",
    "meadow": "meadow",
    "field": "field",
    "jungle": "jungle",
}

URBAN_TYPES: Dict[str, str] = {
    "city": "city",
    "street": "street",
    "road": "street",
    "highway": "street",
    "traffic": "street",
    "market": "market",
    "square": "square",
    "subway": "subway",
    "metro": "subway",
    "station": "station",
    "factory": "industrial",
    "industrial": "industrial",
    "harbor": "harbor",
    "harbour": "harbor",
    "port": "harbor",
}

INTERIOR_TYPES: Dict[str, str] = {
    "kitchen": "household",
    "bedroom": "household",
    "bathroom": "household",
    "living": "household",  # living room will co-occur
    "fireplace": "household",
    "hearth": "household",
    "office": "office",
    "library": "library",
    "bar": "bar",
    "pub": "bar",
    "restaurant": "restaurant",
    "tavern": "bar",
    "inn": "household",
    "hall": "hall",
}

HISTORICAL_TYPES: Dict[str, str] = {
    "castle": "castle",
    "dungeon": "dungeon",
    "ruins": "ruins",
    "palace": "palace",
    "fortress": "castle",
}

SCIFI_TYPES: Dict[str, str] = {
    "spaceship": "spacecraft",
    "starship": "spacecraft",
    "spacecraft": "spacecraft",
    "space": "space",
    "futuristic": "futuristic",
}

COMBAT_TYPES: Dict[str, str] = {
    "battle": "combat",
    "combat": "combat",
    "fight": "combat",
    "fighting": "combat",
    "warfare": "combat",
    "military": "military",
}

# Words that should be treated as locations only (never as Types)
# This list can be extended over time if needed.
LOCATION_ONLY_HINTS: Set[str] = {
    "newcastle",
    "taipei",
    "taiwan",
    "australia",
    "france",
    "spain",
    "london",
    "paris",
    "prague",
    "rome",
    "berlin",
    "tokyo",
    "village",  # handled specially; see below
}

# Texture lexicon (how it sounds)
TEXTURE_WORDS: Dict[str, str] = {
    "boom": "impact",
    "crash": "impact",
    "thud": "impact",
    "slam": "impact",
    "impact": "impact",
    "drone": "drone",
    "droning": "drone",
    "hum": "hum",
    "humming": "hum",
    "buzz": "buzz",
    "buzzing": "buzz",
    "rumble": "rumble",
    "rumbling": "rumble",
    "hiss": "hiss",
    "hissing": "hiss",
    "crackle": "crackle",
    "crackling": "crackle",
    "whoosh": "whoosh",
    "swoosh": "whoosh",
    "rush": "rush",
    "flyby": "flyby",
    "riser": "riser",
    "swell": "swell",
    "stinger": "stinger",
    "cinematic": "cinematic",
    "rhythmic": "rhythmic",
    "percussive": "percussive",
    "pounding": "pounding",
    "glitched": "glitched",
    "resonance": "resonance",
    "spacious": "spacious",
    "heavy": "heavy",
    "light": "light",
}

# One-word SFX type keywords that should live in Type for 03_SFX assets
SFX_TYPE_KEYWORDS: List[str] = [
    "drone",
    "braam",
    "chime",
    "hit",
    "glitch",
]

# Genre lexicon (canonical, one-word labels)

# Canonical genre labels (musical) – exactly as specified by the user
MUSICAL_GENRES: Set[str] = {
    "pop",
    "rock",
    "jazz",
    "classical",
    "orchestral",
    "country",
    "new age",
    "relaxing",
    "world",
    "folk",
    "electronic",
    "celtic",
}

# Canonical genre labels (literary / content) – exactly as specified
LITERARY_GENRES: Set[str] = {
    "sci-fi",
    "fantasy",
    "thriller",
    "mystery",
    "drama",
    "comedy",
    "horror",
    "crime",
    "anime",
    "children",
    "documentary",
    "document",
    "biography",
    "romance",
    "dystopian",
    "adventure",
    "young adult",
    "historical fiction",
    "contemporary fiction",
    "memoir",
    "poetry",
}

ALL_GENRES: Set[str] = MUSICAL_GENRES | LITERARY_GENRES

# Synonym map: lowercased phrase -> canonical genre
GENRE_SYNONYMS: Dict[str, str] = {
    # Musical
    "pop": "pop",
    "rock": "rock",
    "jazz": "jazz",
    "classical": "classical",
    "orchestral": "orchestral",
    "orchestra": "orchestral",
    "country": "country",
    "new age": "new age",
    "new-age": "new age",
    "new_age": "new age",
    "relaxing": "relaxing",
    "relaxation": "relaxing",
    "calm": "relaxing",
    "soothing": "relaxing",
    "ambient": "relaxing",
    "world": "world",
    "folk": "folk",
    "electronic": "electronic",
    "edm": "electronic",
    "synth": "electronic",
    "synthwave": "electronic",
    "electro": "electronic",
    "celtic": "celtic",

    # Literary / content
    "sci-fi": "sci-fi",
    "scifi": "sci-fi",
    "sci fi": "sci-fi",
    "science fiction": "sci-fi",
    "science_fiction": "sci-fi",
    "sciencefiction": "sci-fi",
    "science": "sci-fi",
    "fantasy": "fantasy",
    "thriller": "thriller",
    "mystery": "mystery",
    "detective": "mystery",
    "drama": "drama",
    "comedy": "comedy",
    "humor": "comedy",
    "humour": "comedy",
    "horror": "horror",
    "crime": "crime",
    "noir": "crime",
    "anime": "anime",
    "manga": "anime",
    "cartoon": "anime",
    "children": "children",
    "kids": "children",
    "childrens": "children",
    "kid": "children",
    "documentary": "documentary",
    "document": "document",
    "doc": "documentary",
    "biography": "biography",
    "bio": "biography",
    "romance": "romance",
    "romantic": "romance",
    "dystopian": "dystopian",
    "adventure": "adventure",
    "young adult": "young adult",
    "young_adult": "young adult",
    "ya": "young adult",
    "historical fiction": "historical fiction",
    "historical_fiction": "historical fiction",
    "contemporary fiction": "contemporary fiction",
    "contemporary_fiction": "contemporary fiction",
    "memoir": "memoir",
    "poetry": "poetry",
}

# Era cues
ERA_PREHISTORY: Set[str] = {"dinosaur", "dinosaurs", "prehistoric"}
ERA_ANTIQUITY: Set[str] = {"roman", "rome", "greek", "ancient", "egypt", "egyptian"}
ERA_MEDIEVAL: Set[str] = {"medieval", "castle", "knight", "knights"}
ERA_MODERNITY: Set[str] = {"beethoven", "mozart", "symphony", "orchestra"}
ERA_FUTURE: Set[str] = {"spaceship", "starship", "laser", "cyber", "futuristic", "spacecraft"}

TECHNOLOGY_HINTS: Set[str] = {
    "engine",
    "car",
    "cars",
    "truck",
    "trucks",
    "traffic",
    "train",
    "trains",
    "subway",
    "metro",
    "plane",
    "airplane",
    "helicopter",
    "phone",
    "radio",
    "computer",
}

# Mood patterns
COMBAT_WORDS: Set[str] = {
    "battle",
    "combat",
    "fight",
    "fighting",
    "sword",
    "swords",
    "gun",
    "guns",
    "rifle",
    "shot",
    "shots",
    "explosion",
    "explosions",
    "grenade",
    "bomb",
    "attack",
    "attacks",
}

BLOODY_HINTS: Set[str] = {"blood", "gore", "scream", "screams", "screaming", "body", "bodies"}
THREAT_ONLY_HINTS: Set[str] = {"arrow", "arrows", "distant", "warning"}

HOUSEHOLD_COMFORT: Dict[str, Tuple[str, str]] = {
    "fireplace": ("calm", "cosy"),
    "hearth": ("calm", "cosy"),
    "bathtub": ("relaxed", "intimate"),
    "bath": ("relaxed", "intimate"),
    "bedroom": ("intimate", "calm"),
    "kitchen": ("busy", "warm"),
    "living": ("calm", "cosy"),
}

NATURE_PEACEFUL_HINTS: Set[str] = {"calm", "gentle", "soft", "serene", "peaceful", "quiet"}
NATURE_DRAMATIC_HINTS: Set[str] = {"storm", "thunder", "avalanche", "blizzard", "hurricane", "roaring"}

DARK_ATMOS_LABELS: Set[str] = {"dark_harmonics", "dark", "atmospheres", "atmosphere"}


def choose_era(tokens: Set[str], type_value: Optional[str]) -> Optional[str]:
    """Infer Era following project rules.

    - Nature types -> era-agnostic (None) unless very explicit cues.
    - Technology/urban -> contemporary by default.
    - Very old/future cues mapped to specific era categories.
    - Otherwise, if clearly in the past but ambiguous, use "historical".
    """

    if not tokens:
        return None

    # Nature: no era unless extremely explicit (dinosaurs, jurassic etc.)
    if type_value in {"forest", "jungle", "desert", "beach", "ocean", "sea", "river", "lake", "waterfall", "meadow", "field"}:
        if tokens & ERA_PREHISTORY:
            return "prehistory"
        # otherwise era-agnostic
        return None

    if tokens & ERA_PREHISTORY:
        return "prehistory"
    if tokens & ERA_ANTIQUITY:
        return "antiquity"
    if tokens & ERA_MEDIEVAL:
        # If clearly medieval, use that; otherwise will fall back to historical
        return "medieval"
    if tokens & ERA_MODERNITY:
        return "modernity"
    if tokens & ERA_FUTURE:
        return "future"

    # Technology or urban cues -> contemporary (unless overridden above)
    if tokens & TECHNOLOGY_HINTS or (type_value in {"city", "street", "industrial", "station", "subway"}):
        return "contemporary"

    # If we see strong historical archetypes but not a single era bucket,
    # use generic "historical".
    if tokens & {"castle", "knight", "sword", "battle", "medieval", "ancient"}:
        return "historical"

    return None


def infer_type(tokens: List[str], token_set: Set[str], existing_type: Optional[str], existing_location: Optional[str]) -> Optional[str]:
    """Infer Type (environment/setting) if currently empty.

    Respects:
    - Does not overwrite non-empty existing_type.
    - Uses whole-word matching.
    - Never uses city/location tokens as Type.
    - "village" is never mapped to "city"; it's treated as rural/town.
    - "war" is not used as Type at all.
    """

    if existing_type:
        # Keep manual or legacy value; we only fix very narrow bad cases.
        # Remove forbidden "war" type if it somehow slipped in.
        if existing_type.strip().lower() == "war":
            return None
        return existing_type

    if not tokens:
        return None

    # Quick guards: never use "war" as type
    if "war" in token_set:
        return None

    # Nature
    for t in tokens:
        if t in LOCATION_ONLY_HINTS:
            continue
        if t in NATURE_TYPES:
            return NATURE_TYPES[t]

    # Urban
    for t in tokens:
        if t in LOCATION_ONLY_HINTS:
            continue
        if t in URBAN_TYPES:
            return URBAN_TYPES[t]

    # Interior
    for t in tokens:
        if t in INTERIOR_TYPES:
            return INTERIOR_TYPES[t]

    # Historical structures
    for t in tokens:
        if t in HISTORICAL_TYPES:
            return HISTORICAL_TYPES[t]

    # Sci-fi
    for t in tokens:
        if t in SCIFI_TYPES:
            return SCIFI_TYPES[t]

    # Combat / military
    for t in tokens:
        if t in COMBAT_TYPES:
            return COMBAT_TYPES[t]

    # Village handling: treat as rural/town but not city; do not force type
    if "village" in token_set:
        if any(tok in NATURE_TYPES for tok in tokens):
            return "nature"
        if any(tok in {"traffic", "street", "road", "cars", "city"} for tok in tokens):
            return "town"
        return "town"

    # Horror as Type only if absolutely nothing else
    if "horror" in token_set:
        return "horror"

    return None


def infer_texture(tokens: List[str], existing_texture: Optional[str]) -> Optional[str]:
    """Infer Texture based on how the sound behaves.

    Never overwrites existing non-empty texture.
    """

    if existing_texture:
        return existing_texture

    textures: List[str] = []
    used: Set[str] = set()
    for t in tokens:
        if t in TECHNICAL_WORDS:
            continue
        if t in TEXTURE_WORDS:
            val = TEXTURE_WORDS[t]
            if val not in used:
                textures.append(val)
                used.add(val)

    if not textures:
        return None

    return ", ".join(textures)


def infer_genre(tokens: List[str], existing_genre: Optional[str], type_value: Optional[str]) -> Optional[str]:
    """Infer Genre (literary/musical style).

    Uses the strict one-word canonical vocabulary defined above.

    - Does not duplicate with Type.
    - Never uses "war" as genre (maps to action/historical instead).
    - Horror is primarily Genre; Type gets horror only as last-resort.
    """

    # If an explicit genre is already set, try to normalize it later in the
    # cleanup pass; for inference we treat it as absent.
    if existing_genre:
        # Clean forbidden "war" if present
        if existing_genre.strip().lower() != "war":
            return existing_genre
        existing_genre = None

    if not tokens:
        return None

    # War handling -> action / historical (mapped to literary action/historical
    # fiction; here we choose "adventure" / "historical_fiction" as closest).
    if "war" in tokens or "warfare" in tokens:
        if "medieval" in tokens or "ancient" in tokens or "castle" in tokens:
            return "historical_fiction"
        return "adventure"

    # Try to find a canonical genre via synonyms
    raw = " ".join(tokens).lower()
    candidates: List[str] = []
    for phrase, canon in GENRE_SYNONYMS.items():
        if phrase in raw:
            if canon not in candidates:
                candidates.append(canon)

    if not candidates:
        return None

    # Prefer not to duplicate Type
    type_lc = type_value.strip().lower() if type_value else ""
    for c in candidates:
        if c != type_lc:
            return c

    # Fallback: first candidate
    return candidates[0] if candidates else None


def infer_location(tokens: List[str], existing_location: Optional[str]) -> Optional[str]:
    """Infer Location where safe.

    We are conservative: we only add simple, clearly geographic hints,
    and never overwrite existing non-empty Location.
    """

    if existing_location:
        return existing_location

    # "village" is explicitly a location, but meaning depends on context.
    # Here we just set the location label itself; semantic interpretation
    # is handled in Type inference.
    if "village" in tokens:
        return "village"

    for t in tokens:
        if t in LOCATION_ONLY_HINTS:
            return t

    return None


def infer_moods(tokens: Set[str], existing_mood1: Optional[str], existing_mood2: Optional[str], genre: Optional[str], type_value: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """Infer Mood1 / Mood2 according to project rules.

    - Combat / fight / battle -> aggressive + brutal/tense.
    - Household / fireplace etc. -> calm + cosy or similar.
    - Horror -> scary + tense.
    - Nature peaceful vs dramatic.
    - Dark atmospheric drones -> scary/tense or mysterious/eerie, but NOT aggressive.
    """

    mood1 = existing_mood1.strip() if existing_mood1 else ""
    mood2 = existing_mood2.strip() if existing_mood2 else ""

    # If both moods present, respect them.
    if mood1 and mood2:
        return mood1, mood2

    # Normalize genre for checks
    genre_lc = genre.lower() if genre else ""
    type_lc = type_value.lower() if type_value else ""

    # Combat moods
    if tokens & COMBAT_WORDS:
        if not mood1:
            mood1 = "aggressive"
        # Brutal vs tense depending on presence of damage/blood/body
        if not mood2:
            if tokens & BLOODY_HINTS:
                mood2 = "brutal"
            else:
                mood2 = "tense"

    # Household comfort
    for word, (m1, m2) in HOUSEHOLD_COMFORT.items():
        if word in tokens:
            if not mood1:
                mood1 = m1
            if not mood2:
                mood2 = m2
            break

    # Horror genre
    if "horror" in tokens or genre_lc == "horror":
        if not mood1:
            mood1 = "scary"
        if not mood2:
            mood2 = "tense"

    # Genre-driven moods (LLM-style semantic inference, but rule-based and
    # anchored strictly in the known Genre label – no free-form guessing).

    # Suspense / mystery / crime / dystopian / sci-fi
    if not (mood1 and mood2):
        if genre_lc in {"thriller", "mystery", "crime", "dystopian", "sci-fi"}:
            if not mood1:
                mood1 = "tense"
            if not mood2:
                mood2 = "mysterious"

    # Drama / serious narrative
    if not (mood1 and mood2):
        if genre_lc in {"drama", "historical fiction", "contemporary fiction", "memoir"}:
            if not mood1:
                mood1 = "emotional"
            if not mood2:
                mood2 = "dramatic"

    # Romance
    if not (mood1 and mood2):
        if genre_lc == "romance":
            if not mood1:
                mood1 = "romantic"
            if not mood2:
                mood2 = "tender"

    # Comedy / light-hearted
    if not (mood1 and mood2):
        if genre_lc in {"comedy", "children", "anime"}:
            if not mood1:
                mood1 = "funny"
            if not mood2:
                mood2 = "playful"

    # Documentary / neutral
    if not (mood1 and mood2):
        if genre_lc in {"documentary", "document", "biography"}:
            if not mood1:
                mood1 = "neutral"
            if not mood2:
                mood2 = "informative"

    # Fantasy / adventure
    if not (mood1 and mood2):
        if genre_lc in {"fantasy", "adventure"}:
            if not mood1:
                mood1 = "adventurous"
            if not mood2:
                mood2 = "epic"

    # Sci-fi with strong technology cues but no other moods yet
    if not (mood1 and mood2):
        if genre_lc == "sci-fi" or (genre_lc == "" and tokens & ERA_FUTURE):
            if not mood1:
                mood1 = "futuristic"
            if not mood2:
                mood2 = "tense"

    # Nature moods
    if type_lc in {"forest", "jungle", "desert", "beach", "ocean", "sea", "river", "lake", "waterfall", "meadow", "field", "nature"}:
        if tokens & NATURE_PEACEFUL_HINTS:
            if not mood1:
                mood1 = "serene"
            if not mood2:
                mood2 = "peaceful"
        elif tokens & NATURE_DRAMATIC_HINTS:
            if not mood1:
                mood1 = "majestic"
            if not mood2:
                mood2 = "powerful"

    # Dark atmospheric drones (e.g. Dark_Harmonics)
    if tokens & DARK_ATMOS_LABELS or ("dark" in tokens and ("atmosphere" in tokens or "atmospheric" in tokens)):
        if not mood1:
            mood1 = "mysterious"
        if not mood2:
            mood2 = "tense"

    return mood1 or None, mood2 or None


def _split_multi_value(value: Optional[str]) -> List[str]:
    """Split a multi-value cell (comma-separated) into normalized tokens."""

    if not value:
        return []
    parts: List[str] = []
    for chunk in str(value).split(","):
        token = chunk.strip()
        if not token:
            continue
        parts.append(token)
    return parts


def _normalise_mood_term(term: str) -> str:
    """Normalise individual mood term according to explicit mappings.

    Mappings (case-insensitive):
    - scared  -> scary
    - tension -> tense
    - mystery -> mysterious
    - spooky  -> eerie
    - fearful -> scary
    """
    value = term.strip()
    if not value:
        return value
    l = value.lower()
    if l == "scared":
        return "scary"
    if l == "tension":
        return "tense"
    if l == "mystery":
        return "mysterious"
    if l == "spooky":
        return "eerie"
    if l == "fearful":
        return "scary"
    return value


def _collect_row_words(row_values: Dict[str, Optional[str]]) -> Set[str]:
    """Collect all existing lowercased words across key semantic fields.

    This prevents duplicates across Type/Texture/Genre/Era/Location/Moods
    when we *append* metadata-derived words.
    """

    words: Set[str] = set()
    for val in row_values.values():
        for part in _split_multi_value(val):
            words.add(part.lower())
    return words


def _merge_values(existing: Optional[str], candidate: Optional[str], row_words: Set[str]) -> Tuple[Optional[str], Set[str]]:
    """Merge candidate value into an existing comma-separated cell.

    - Never removes or rewrites existing content (filename-derived words).
    - Appends *new* tokens separated by ", ".
    - Skips tokens already present anywhere in the row (row_words) to
      respect the "NO DUPLICATES" rule (1 mined word -> 1 field).
    """

    if not candidate:
        return existing, row_words

    existing_parts = _split_multi_value(existing)
    parts_lower = {p.lower() for p in existing_parts}

    added_any = False
    for token in _split_multi_value(candidate):
        low = token.lower()
        if low in row_words or low in parts_lower:
            continue
        existing_parts.append(token)
        parts_lower.add(low)
        row_words.add(low)
        added_any = True

    if not existing_parts and not added_any:
        return existing, row_words

    merged = ", ".join(existing_parts)
    return merged, row_words


def normalize_genre_cell(value: Optional[str], group: str) -> Optional[str]:
    """Normalize a Genre cell to a single canonical label from the allowed set.

    Rules:
    - Input may contain multiple comma-separated items; we normalise each part.
    - Result is always at most ONE canonical label (no commas).
    - Only values from MUSICAL_GENRES/LITERARY_GENRES are allowed.
    - If no clear mapping exists for any part, we clear the cell (return None).
    - For 02_MUSIC rows, prefer musical genres when multiple candidates match.
    - For other rows, prefer literary genres when possible.
    """

    if not value:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    # Split on commas and normalise each chunk independently
    parts = _split_multi_value(raw)
    if not parts:
        return None

    canonical_candidates: List[str] = []
    for part in parts:
        l = part.strip().lower()
        if not l:
            continue
        # Direct canonical match
        for canon in ALL_GENRES:
            if l == canon.lower():
                if canon not in canonical_candidates:
                    canonical_candidates.append(canon)
                break
        else:
            # Try synonym mapping
            for phrase, canon in GENRE_SYNONYMS.items():
                if phrase in l:
                    if canon not in canonical_candidates:
                        canonical_candidates.append(canon)
                    break

    if not canonical_candidates:
        # Nothing mapped to our controlled vocabulary -> clear the cell
        return None

    is_music = group == "02_MUSIC"
    preferred_pool = MUSICAL_GENRES if is_music else LITERARY_GENRES
    secondary_pool = LITERARY_GENRES if is_music else MUSICAL_GENRES

    # First try to pick from the preferred pool
    for c in canonical_candidates:
        if c in preferred_pool:
            return c

    # Then fall back to the secondary pool
    for c in canonical_candidates:
        if c in secondary_pool:
            return c

    # As a last resort, pick the first canonical candidate
    return canonical_candidates[0]


def process_excel(excel_path: Path, dry_run: bool = False) -> Dict[str, int]:
    """Apply inference rules to catalog.xlsx.

    Returns a dict of counters for reporting.
    """

    wb = load_workbook(excel_path)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{EXCEL_SHEET_NAME}' not found in {excel_path}")

    ws = wb[EXCEL_SHEET_NAME]

    # Build header -> column index mapping
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = idx

    required = [
        "AssetID",
        "Filename",
        "Type",
        "Texture",
        "Genre",
        "Era",
        "Location",
        "Source1",
        "Source2",
        "Source3",
        "Mood1",
        "Mood2",
        "Intensity",
        # NOTE: "Duration" is intentionally NOT required here. The Excel
        # schema currently does not include it, and per project instructions
        # we must not add new columns. Duration may be introduced later and
        # is handled as optional in JSON/manifest export.
    ]
    for name in required:
        if name not in headers:
            raise SystemExit(f"Expected column '{name}' not found in sheet '{EXCEL_SHEET_NAME}'")

    stats = {
        "rows_total": 0,
        "type_filled": 0,
        "texture_filled": 0,
        "genre_filled": 0,
        "era_filled": 0,
        "location_filled": 0,
        "mood1_filled": 0,
        "mood2_filled": 0,
    }

    base_dir = excel_path.parent  # soundscape/assets

    for row in ws.iter_rows(min_row=2):
        stats["rows_total"] += 1

        asset_id = row[headers["AssetID"] - 1].value or ""
        filename = row[headers["Filename"] - 1].value or ""

        type_cell = row[headers["Type"] - 1]
        texture_cell = row[headers["Texture"] - 1]
        genre_cell = row[headers["Genre"] - 1]
        era_cell = row[headers["Era"] - 1]
        location_cell = row[headers["Location"] - 1]
        mood1_cell = row[headers["Mood1"] - 1]
        mood2_cell = row[headers["Mood2"] - 1]

        existing_type = str(type_cell.value).strip() if type_cell.value else None
        existing_texture = str(texture_cell.value).strip() if texture_cell.value else None
        existing_genre = str(genre_cell.value).strip() if genre_cell.value else None
        existing_era = str(era_cell.value).strip() if era_cell.value else None
        existing_location = str(location_cell.value).strip() if location_cell.value else None
        existing_mood1 = str(mood1_cell.value).strip() if mood1_cell.value else None
        existing_mood2 = str(mood2_cell.value).strip() if mood2_cell.value else None

        # Primary tokens from AssetID + Filename (highest priority)
        file_tokens, file_token_set = build_token_context(str(asset_id), str(filename))

        # Secondary tokens from embedded audio tags (if available)
        meta_tokens, meta_token_set = extract_metadata_tokens(base_dir, str(asset_id), str(filename))

        all_tokens = list(file_tokens)
        for t in meta_tokens:
            if t not in file_token_set:
                all_tokens.append(t)
        all_token_set: Set[str] = set(all_tokens)

        # Infer fields using combined context; filename-derived content always
        # wins because we never overwrite existing non-empty cells.
        type_val = infer_type(all_tokens, all_token_set, existing_type, existing_location)
        texture_val = infer_texture(all_tokens, existing_texture)
        genre_val = infer_genre(all_tokens, existing_genre, type_val)
        era_val = choose_era(all_token_set, type_val or existing_type)
        location_val = infer_location(all_tokens, existing_location)
        mood1_val, mood2_val = infer_moods(all_token_set, existing_mood1, existing_mood2, genre_val or existing_genre, type_val or existing_type)

        # Apply changes only if cell was empty and we inferred something
        if not existing_type and type_val:
            stats["type_filled"] += 1
            if not dry_run:
                type_cell.value = type_val

        if not existing_texture and texture_val:
            stats["texture_filled"] += 1
            if not dry_run:
                texture_cell.value = texture_val

        if not existing_genre and genre_val:
            stats["genre_filled"] += 1
            if not dry_run:
                genre_cell.value = genre_val

        if not existing_era and era_val:
            stats["era_filled"] += 1
            if not dry_run:
                era_cell.value = era_val

        if not existing_location and location_val:
            stats["location_filled"] += 1
            if not dry_run:
                location_cell.value = location_val

        if not existing_mood1 and mood1_val:
            stats["mood1_filled"] += 1
            if not dry_run:
                mood1_cell.value = mood1_val

        if not existing_mood2 and mood2_val:
            stats["mood2_filled"] += 1
            if not dry_run:
                mood2_cell.value = mood2_val

        # --- Second-pass: augment from metadata tokens only -----------------
        # At this point, filename-derived content is already in the cells.
        # We now *append* useful words from metadata, without ever
        # overwriting or duplicating existing ones.

        if meta_tokens and not dry_run:
            row_values = {
                "Type": type_cell.value if type_cell.value is not None else None,
                "Texture": texture_cell.value if texture_cell.value is not None else None,
                "Genre": genre_cell.value if genre_cell.value is not None else None,
                "Era": era_cell.value if era_cell.value is not None else None,
                "Location": location_cell.value if location_cell.value is not None else None,
                "Mood1": mood1_cell.value if mood1_cell.value is not None else None,
                "Mood2": mood2_cell.value if mood2_cell.value is not None else None,
            }

            row_words = _collect_row_words(row_values)

            meta_tokens_list = list(meta_token_set)

            # Derive candidate values from metadata in isolation
            m_type = infer_type(meta_tokens_list, meta_token_set, None, None)
            m_texture = infer_texture(meta_tokens_list, None)
            m_genre = infer_genre(meta_tokens_list, None, m_type)
            m_era = choose_era(meta_token_set, m_type or existing_type)
            m_location = infer_location(meta_tokens_list, None)
            m_mood1, m_mood2 = infer_moods(meta_token_set, None, None, m_genre or existing_genre, m_type or existing_type)

            # Merge into existing cells, appending non-duplicate tokens.
            new_type, row_words = _merge_values(row_values["Type"], m_type, row_words)
            type_cell.value = new_type

            new_texture, row_words = _merge_values(row_values["Texture"], m_texture, row_words)
            texture_cell.value = new_texture

            new_genre, row_words = _merge_values(row_values["Genre"], m_genre, row_words)
            genre_cell.value = new_genre

            new_era, row_words = _merge_values(row_values["Era"], m_era, row_words)
            era_cell.value = new_era

            new_loc, row_words = _merge_values(row_values["Location"], m_location, row_words)
            location_cell.value = new_loc

            new_m1, row_words = _merge_values(row_values["Mood1"], m_mood1, row_words)
            mood1_cell.value = new_m1

            new_m2, row_words = _merge_values(row_values["Mood2"], m_mood2, row_words)
            mood2_cell.value = new_m2

        # --- Explicit folder / keyword rules --------------------------------
        # Apply after generic inference so that these project-specific rules
        # can override or fill in where needed.

        if not dry_run:
            rel = str(asset_id).replace("\\", "/")
            segments = [s for s in rel.split("/") if s]
            group = segments[0] if segments else ""
            sub1 = segments[1] if len(segments) > 1 else ""

            # 01_AMB: if Type is still empty, use its first subfolder name
            if group == "01_AMB":
                if not (type_cell.value and str(type_cell.value).strip()) and sub1:
                    type_cell.value = sub1.lower()

            # 02_MUSIC: all files use Type "music" (override if needed)
            if group == "02_MUSIC":
                type_cell.value = "music"

            # Prepare filename tokens for 03_SFX rules
            base_name = str(filename).rsplit(".", 1)[0]
            fname_tokens = tokenize(base_name)
            fname_token_set = set(fname_tokens)

            if group == "03_SFX":
                # All animals SFX -> Type: animal
                if any(seg.lower() == "animals" for seg in segments):
                    type_cell.value = "animal"

                # All alarm and siren SFX -> Era: contemporary, Type: emergency
                if "alarm" in fname_token_set or "siren" in fname_token_set:
                    type_cell.value = "emergency"
                    era_cell.value = "contemporary"

                # Cartoon/anime -> Type: cartoon, Genre: anime (single word)
                if any("cartoon" in seg.lower() or "anime" in seg.lower() for seg in segments):
                    type_cell.value = "cartoon"
                    genre_cell.value = "anime"

                # SFX keyword types: use specific one-word types like drone,
                # braam, chime, hit, glitch for Type. These should not be
                # duplicated in Texture/Genre/Moods.
                chosen_kw: Optional[str] = None
                for kw in SFX_TYPE_KEYWORDS:
                    if kw in fname_token_set:
                        chosen_kw = kw
                        break

                # Only apply keyword type if this SFX is not already forced
                # into more specific categories above (animal, emergency,
                # cartoon) to avoid overriding those explicit rules.
                if chosen_kw and str(type_cell.value or "").lower() not in {"animal", "emergency", "cartoon"}:
                    # Remove keyword from other semantic fields to avoid
                    # duplicates, keeping 1 word -> 1 field.
                    for cell in (texture_cell, genre_cell, mood1_cell, mood2_cell):
                        val = cell.value
                        if not val:
                            continue
                        parts = _split_multi_value(str(val))
                        parts = [p for p in parts if p.strip().lower() != chosen_kw]
                        cell.value = ", ".join(parts) if parts else ""

                    type_cell.value = chosen_kw

        # --- Normalize Genre to canonical labels ----------------------------
        # This runs for all rows and ensures Genre is a single canonical
        # value from the controlled vocabulary when a clear mapping exists.

        if not dry_run:
            rel = str(asset_id).replace("\\", "/")
            segments = [s for s in rel.split("/") if s]
            group = segments[0] if segments else ""

            if genre_cell.value is not None and str(genre_cell.value).strip():
                normalized = normalize_genre_cell(str(genre_cell.value), group)
                genre_cell.value = normalized

            # --- Clean up moods --------------------------------------------
            # 1) "cinematic" is not a mood: move it into Texture, preserving
            #    existing Texture values (comma-separated) and removing it
            #    from Mood1/Mood2.
            # 2) Apply explicit term replacements in moods:
            #    scared->scary, tension->tense, mystery->mysterious,
            #    spooky->eerie, fearful->scary.

            tex_val = str(texture_cell.value).strip() if texture_cell.value else ""
            m1_val = str(mood1_cell.value).strip() if mood1_cell.value else ""
            m2_val = str(mood2_cell.value).strip() if mood2_cell.value else ""

            tex_parts = _split_multi_value(tex_val)
            m1_parts = _split_multi_value(m1_val)
            m2_parts = _split_multi_value(m2_val)

            tex_lower = {p.lower() for p in tex_parts}

            # Move cinematic from moods to texture
            filtered_m1: List[str] = []
            for p in m1_parts:
                if p.strip().lower() == "cinematic":
                    if "cinematic" not in tex_lower:
                        tex_parts.append("cinematic")
                        tex_lower.add("cinematic")
                else:
                    filtered_m1.append(p)

            filtered_m2: List[str] = []
            for p in m2_parts:
                if p.strip().lower() == "cinematic":
                    if "cinematic" not in tex_lower:
                        tex_parts.append("cinematic")
                        tex_lower.add("cinematic")
                else:
                    filtered_m2.append(p)

            # Apply mood term normalisation
            norm_m1 = [_normalise_mood_term(p) for p in filtered_m1]
            norm_m2 = [_normalise_mood_term(p) for p in filtered_m2]

            # Deduplicate within each mood cell (case-insensitive)
            def _dedup(parts: List[str]) -> List[str]:
                seen: Set[str] = set()
                out: List[str] = []
                for p in parts:
                    lp = p.lower()
                    if lp in seen:
                        continue
                    seen.add(lp)
                    out.append(p)
                return out

            norm_m1 = _dedup(norm_m1)
            norm_m2 = _dedup(norm_m2)

            texture_cell.value = ", ".join(tex_parts) if tex_parts else ""
            mood1_cell.value = ", ".join(norm_m1) if norm_m1 else ""
            mood2_cell.value = ", ".join(norm_m2) if norm_m2 else ""

    if not dry_run:
        wb.save(excel_path)

    return stats


def export_catalog_json(excel_path: Path, json_path: Path) -> None:
    """Export Excel sheet to catalog.json in current format.

    The JSON is an array of objects, each mirroring one row.
    """

    wb = load_workbook(excel_path, read_only=True)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{EXCEL_SHEET_NAME}' not found in {excel_path}")

    ws = wb[EXCEL_SHEET_NAME]

    headers: List[str] = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    items: List[Dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # skip fully empty rows
            continue
        obj: Dict[str, object] = {}
        for idx, value in enumerate(row):
            if idx >= len(headers):
                continue
            key = headers[idx]
            if not key:
                continue
            obj[key] = value if value is not None else ""
        items.append(obj)

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def export_manifest_json(excel_path: Path, manifest_path: Path) -> None:
    """Export Excel sheet to manifest.json in current format.

    Structure:
    {
      "version": "1.0",
      "generated": "generate_catalog_v2.py",
      "assetCount": N,
      "assets": {
        "AssetID": {"filename": "...", "duration": "..."},
        ...
      }
    }
    """

    wb = load_workbook(excel_path, read_only=True)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{EXCEL_SHEET_NAME}' not found in {excel_path}")

    ws = wb[EXCEL_SHEET_NAME]

    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = idx

    # We require AssetID and Filename only; Duration is optional and
    # not present in the current Excel schema. When missing, we simply
    # emit an empty string in the manifest.
    for key in ("AssetID", "Filename"):
        if key not in headers:
            raise SystemExit(f"Expected column '{key}' not found in sheet '{EXCEL_SHEET_NAME}'")

    assets: Dict[str, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2):
        asset_id = row[headers["AssetID"] - 1].value
        if not asset_id:
            continue
        filename = row[headers["Filename"] - 1].value or ""

        duration = ""
        duration_idx = headers.get("Duration")
        if duration_idx is not None:
            duration_val = row[duration_idx - 1].value
            if duration_val is not None:
                duration = str(duration_val)
        assets[str(asset_id)] = {
            "filename": str(filename),
            "duration": duration,
        }

    manifest = {
        "version": "1.0",
        "generated": "generate_catalog_v2.py",
        "assetCount": len(assets),
        "assets": assets,
    }

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)


def main(argv: Optional[Sequence[str]] = None) -> None:
    parser = argparse.ArgumentParser(description="Generate/enrich soundscape catalog.xlsx (Phase 2)")
    parser.add_argument("--excel", type=Path, default=Path("soundscape/assets/catalog.xlsx"), help="Path to catalog.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Analyze and report, but do not modify Excel")
    parser.add_argument("--export-json", action="store_true", help="Also export updated Excel to catalog.json")
    parser.add_argument("--export-manifest", action="store_true", help="Also export manifest.json from updated Excel")

    args = parser.parse_args(list(argv) if argv is not None else None)

    excel_path: Path = args.excel
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    stats = process_excel(excel_path, dry_run=args.dry_run)

    rows = stats["rows_total"] or 1
    print("=== Excel Enrichment Summary ===")
    print(f"Rows total: {stats['rows_total']}")
    print(f"Type filled:     {stats['type_filled']}  ({stats['type_filled']/rows*100:.1f}% of rows)")
    print(f"Texture filled:  {stats['texture_filled']}  ({stats['texture_filled']/rows*100:.1f}% of rows)")
    print(f"Genre filled:    {stats['genre_filled']}  ({stats['genre_filled']/rows*100:.1f}% of rows)")
    print(f"Era filled:      {stats['era_filled']}  ({stats['era_filled']/rows*100:.1f}% of rows)")
    print(f"Location filled: {stats['location_filled']}  ({stats['location_filled']/rows*100:.1f}% of rows)")
    print(f"Mood1 filled:    {stats['mood1_filled']}  ({stats['mood1_filled']/rows*100:.1f}% of rows)")
    print(f"Mood2 filled:    {stats['mood2_filled']}  ({stats['mood2_filled']/rows*100:.1f}% of rows)")

    # Only export JSON/manifest when explicitly requested.
    if not args.dry_run and args.export_json:
        json_path = excel_path.with_name("catalog.json")
        print(f"Writing catalog JSON to {json_path} ...")
        export_catalog_json(excel_path, json_path)

    if not args.dry_run and args.export_manifest:
        manifest_path = excel_path.with_name("manifest.json")
        print(f"Writing manifest JSON to {manifest_path} ...")
        export_manifest_json(excel_path, manifest_path)


if __name__ == "__main__":  # pragma: no cover
    main()

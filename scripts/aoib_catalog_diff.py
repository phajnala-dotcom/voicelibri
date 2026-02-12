import csv
import datetime as _dt
import os
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import quote


def _long_path(path: Path) -> str:
    s = str(path)
    if os.name == 'nt' and not s.startswith('\\\\?\\'):
        return '\\\\?\\' + s
    return s


def _strip_long_prefix(p: str) -> str:
    return p[4:] if p.startswith('\\\\?\\') else p


@dataclass(frozen=True)
class CatalogEntry:
    rec_id: str
    filename: str
    file_path: str
    category: str
    category_full: str


def _norm_slashes(path: str) -> str:
    return path.replace('\\', '/').replace('//', '/')


def _strip_prefix(path: str, prefix: str) -> str:
    if path.lower().startswith(prefix.lower()):
        return path[len(prefix) :]
    return path


def catalog_expected_rel_ogg_path(file_path: str) -> str:
        """Normalize catalog FilePath to a comparable .ogg path.

        Supports both updated paths (realistic/.../file.ogg) and legacy
        AIOB 4824/.../file.wav paths.
        """

        p = _norm_slashes(file_path).strip().strip('"')
        if p.lower().endswith('.wav'):
                p = p[: -4] + '.ogg'
        return p


def read_catalog_csv(csv_path: Path) -> list[CatalogEntry]:
    entries: list[CatalogEntry] = []
    with csv_path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            entries.append(
                CatalogEntry(
                    rec_id=str(row.get('RecID', '')).strip(),
                    filename=str(row.get('Filename', '')).strip(),
                    file_path=str(row.get('FilePath', '')).strip(),
                    category=str(row.get('Category', '')).strip(),
                    category_full=str(row.get('CategoryFull', '')).strip(),
                )
            )
    return entries


def read_catalog_xlsx(xlsx_path: Path) -> list[CatalogEntry]:
    # Official docs compliance: use openpyxl (local file read).
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        # Find header row (look for FilePath)
        header_row = None
        header: list[str] = []
        for r in range(1, 31):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            row = [str(v).strip() if v is not None else '' for v in vals]
            if 'FilePath' in row:
                header_row = r
                header = row
                break
        if header_row is None:
            raise RuntimeError('Could not find header row containing FilePath')

        def col(name: str) -> int | None:
            try:
                return header.index(name) + 1
            except ValueError:
                return None

        # Support both original (RecID) and optimized (FileID) catalogs.
        recid_col = col('RecID') or col('FileID')
        filename_col = col('Filename')
        file_path_col = col('FilePath')
        category_col = col('Category')
        category_full_col = col('CategoryFull')

        if file_path_col is None:
            raise RuntimeError('Catalog XLSX missing FilePath column')

        entries: list[CatalogEntry] = []
        for r in range(header_row + 1, ws.max_row + 1):
            fp = ws.cell(row=r, column=file_path_col).value
            if fp is None or str(fp).strip() == '':
                continue
            entries.append(
                CatalogEntry(
                    rec_id=str(ws.cell(row=r, column=recid_col).value).strip() if recid_col else '',
                    filename=str(ws.cell(row=r, column=filename_col).value).strip() if filename_col else '',
                    file_path=str(fp).strip(),
                    category=str(ws.cell(row=r, column=category_col).value).strip() if category_col else '',
                    category_full=str(ws.cell(row=r, column=category_full_col).value).strip() if category_full_col else '',
                )
            )
        return entries
    finally:
        wb.close()


def read_catalog(path: Path) -> list[CatalogEntry]:
    if path.suffix.lower() == '.csv':
        return read_catalog_csv(path)
    if path.suffix.lower() == '.xlsx':
        return read_catalog_xlsx(path)
    raise ValueError(f'Unsupported catalog type: {path.suffix}')


def ensure_catalog_csv(xlsx_path: Path, csv_path: Path) -> Path:
    if csv_path.exists() and csv_path.stat().st_mtime >= xlsx_path.stat().st_mtime:
        return csv_path
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header: list[str] = []
    header_row = 1
    for r in range(1, 10):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_s = [str(v).strip() if v is not None else '' for v in vals]
        if 'FilePath' in row_s:
            header = row_s
            header_row = r
            break

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open('w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            writer.writerow(row)

    wb.close()
    return csv_path


def iter_ogg_files(roots: dict[str, Path]) -> list[str]:
    # Long-path safe inventory (pathlib/glob can miss >260 char paths on Windows).
    out: list[str] = []
    seen: set[str] = set()
    for root_name, root in roots.items():
        root_lp = _long_path(root)
        root_str = _strip_long_prefix(str(root))
        for dirpath, _dirs, files in os.walk(root_lp):
            for fname in files:
                if not fname.lower().endswith('.ogg'):
                    continue
                full = os.path.join(dirpath, fname)
                key = full.lower()
                if key in seen:
                    continue
                seen.add(key)
                rel = _norm_slashes(os.path.relpath(_strip_long_prefix(full), root_str))
                out.append(f"{root_name}/{rel}")
    out.sort(key=str.lower)
    return out


def _to_md_link(rel_path: str) -> str:
    # Report lives under docs/reports/, so use a correct relative path.
    # Also URL-encode the link target so spaces/commas remain clickable.
    rel = rel_path.replace('\\', '/').lstrip('/')
    target = f"../../soundscape/assets/{rel}"
    return f"[{rel}]({quote(target, safe='/:._-')})"


def write_diff_report(
    *,
    catalog_path: Path,
    roots: dict[str, Path],
    out_md_path: Path,
    max_list: int = 2000,
) -> dict[str, int]:
    print('Reading catalog...', flush=True)
    entries = read_catalog(catalog_path)
    print(f'  Catalog loaded: {len(entries)} entries', flush=True)

    expected_rel: dict[str, CatalogEntry] = {}
    expected_rel_lc: dict[str, str] = {}
    for e in entries:
        if not e.file_path:
            continue
        rel = catalog_expected_rel_ogg_path(e.file_path)
        if rel.startswith('realistic/') or rel.startswith('cinematic_&_foley/') or rel.startswith('music/'):
            expected_rel[rel] = e
            expected_rel_lc[rel.lower()] = rel
            continue
        rel = _strip_prefix(rel, 'AIOB 4824/')
        for root_name in roots.keys():
            key = f"{root_name}/{rel}"
            expected_rel[key] = e
            expected_rel_lc[key.lower()] = key

    print('Scanning OGG files on disk...', flush=True)
    actual_rel = iter_ogg_files(roots)
    print(f'  Found {len(actual_rel)} OGG files', flush=True)
    actual_rel_set = set(actual_rel)
    actual_rel_lc = {p.lower(): p for p in actual_rel}

    expected_set = set(expected_rel.keys())

    missing_exact = sorted(expected_set - actual_rel_set)
    extra_exact = sorted(actual_rel_set - expected_set)

    def _top(rel: str) -> str:
        rel_n = rel.replace('\\', '/').lstrip('/')
        parts = rel_n.split('/')
        if parts and parts[0] in roots:
            return parts[1] if len(parts) > 1 else parts[0]
        return parts[0] if parts else '_root'

    # Category stats (top-level folder under aoib_ogg)
    expected_by_cat: dict[str, int] = {}
    found_by_cat: dict[str, int] = {}
    missing_by_cat: dict[str, int] = {}
    extra_by_cat: dict[str, int] = {}

    for rel in expected_set:
        cat = _top(rel)
        expected_by_cat[cat] = expected_by_cat.get(cat, 0) + 1
        if rel in actual_rel_set:
            found_by_cat[cat] = found_by_cat.get(cat, 0) + 1
        else:
            missing_by_cat[cat] = missing_by_cat.get(cat, 0) + 1

    for rel in extra_exact:
        cat = _top(rel)
        extra_by_cat[cat] = extra_by_cat.get(cat, 0) + 1

    missing_case_only: list[str] = []
    for rel in missing_exact[:]:
        if rel.lower() in actual_rel_lc:
            missing_case_only.append(rel)

    extra_case_only: list[str] = []
    for rel in extra_exact[:]:
        if rel.lower() in expected_rel_lc:
            extra_case_only.append(rel)

    now = _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    coverage_pct = (len(actual_rel_set) / len(expected_set) * 100.0) if expected_set else 0.0
    missing_pct = (len(missing_exact) / len(expected_set) * 100.0) if expected_set else 0.0
    extra_pct_of_found = (len(extra_exact) / len(actual_rel_set) * 100.0) if actual_rel_set else 0.0

    print(f'  Missing: {len(missing_exact)}, Extra: {len(extra_exact)}', flush=True)
    print('Writing report...', flush=True)
    out_md_path.parent.mkdir(parents=True, exist_ok=True)
    with out_md_path.open('w', encoding='utf-8', newline='\n') as out:
        out.write(f"# Audio Library Completeness Report\n\n")
        out.write(f"Generated: {now}\n\n")
        out.write(
            f"**Stats:** Expected **{len(expected_set)}**, Found **{len(actual_rel_set)}** ({coverage_pct:.1f}%), Missing **{len(missing_exact)}** ({missing_pct:.1f}%), Extra **{len(extra_exact)}** ({extra_pct_of_found:.1f}% of found)\n\n"
        )
        out.write(f"Catalog: `{catalog_path}`\n\n")
        out.write(f"Audio roots: `{', '.join(str(p) for p in roots.values())}`\n\n")

        out.write("## Missing by Category (Top-Level Folder)\n\n")
        out.write("Sorted by **missing %** (then missing count). Use this to decide which category ZIPs to re-download/re-convert.\n\n")
        out.write("| Category | Expected | Found | Missing | Missing % | Extra |\n")
        out.write("|---|---:|---:|---:|---:|---:|\n")

        rows = []
        for cat, exp in expected_by_cat.items():
            miss = missing_by_cat.get(cat, 0)
            found = found_by_cat.get(cat, 0)
            extra = extra_by_cat.get(cat, 0)
            miss_pct_cat = (miss / exp * 100.0) if exp else 0.0
            rows.append((miss_pct_cat, miss, cat, exp, found, extra))
        rows.sort(key=lambda r: (-r[0], -r[1], r[2].lower()))

        for miss_pct_cat, miss, cat, exp, found, extra in rows:
            cat_target = quote(f"../../soundscape/assets/{cat}", safe='/:._-')
            cat_link = f"[{cat}]({cat_target})"
            out.write(f"| {cat_link} | {exp} | {found} | {miss} | {miss_pct_cat:.1f}% | {extra} |\n")

        out.write("\n")

        out.write("## Summary\n\n")
        out.write(f"- Catalog entries: **{len(entries)}**\n")
        out.write(f"- Expected OGG paths (unique): **{len(expected_set)}**\n")
        out.write(f"- Actual OGG files found: **{len(actual_rel_set)}**\n")
        out.write(f"- Missing (exact path): **{len(missing_exact)}**\n")
        out.write(f"- Extra (not in catalog): **{len(extra_exact)}**\n")
        out.write(f"- Missing due to case-only mismatch: **{len(missing_case_only)}**\n")
        out.write(f"- Extra due to case-only mismatch: **{len(extra_case_only)}**\n\n")

        out.write("## Missing (expected in catalog but not found on disk)\n\n")
        if missing_exact:
            for rel in missing_exact[:max_list]:
                out.write(f"- {_to_md_link(rel)}\n")
            if len(missing_exact) > max_list:
                out.write(f"\n(Truncated: showing first {max_list} items.)\n")
        else:
            out.write("- None\n")

        out.write("\n## Extra (found on disk but not present in catalog)\n\n")
        if extra_exact:
            for rel in extra_exact[:max_list]:
                out.write(f"- {_to_md_link(rel)}\n")
            if len(extra_exact) > max_list:
                out.write(f"\n(Truncated: showing first {max_list} items.)\n")
        else:
            out.write("- None\n")

        out.write("\n## Repair Guidance\n\n")
        out.write(
            "- For **Missing** files: ensure the file exists under `soundscape/assets/realistic`, `cinematic_&_foley`, or `music` with the exact relative path.\n"
        )
        out.write(
            "- For **Extra** files: either (a) add them to the catalog if they are valid assets, or (b) delete/move them out of the roots if they are accidental artifacts.\n"
        )
        out.write(
            "- For **case-only mismatches**: normalize file/folder casing to match the catalog paths (helps cross-platform tooling and future automation).\n"
        )
        out.write(
            "\nIf you want, you can use `scripts/aoib_bulk_edit/aoib_bulk_edit.py` (scaffolded next) to apply path/rename/delete operations in bulk and update the catalog in lockstep.\n"
        )

    return {
        'catalog_entries': len(entries),
        'expected_unique': len(expected_set),
        'actual_files': len(actual_rel_set),
        'missing': len(missing_exact),
        'extra': len(extra_exact),
        'missing_case_only': len(missing_case_only),
        'extra_case_only': len(extra_case_only),
    }


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    xlsx_path = repo_root / 'soundscape' / 'assets' / 'voicelibri_assets_catalog.xlsx'
    csv_path = repo_root / 'soundscape' / 'assets' / 'voicelibri_assets_catalog.csv'
    catalog_path = ensure_catalog_csv(xlsx_path, csv_path)
    roots = {
        'realistic': repo_root / 'soundscape' / 'assets' / 'realistic',
        'cinematic_&_foley': repo_root / 'soundscape' / 'assets' / 'cinematic_&_foley',
        'music': repo_root / 'soundscape' / 'assets' / 'music',
    }
    ts = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    out_md = repo_root / 'docs' / 'reports' / f'audio_library_completeness_{ts}.md'
    stats = write_diff_report(
        catalog_path=catalog_path, roots=roots, out_md_path=out_md
    )
    print(f"Wrote: {out_md}")
    print(stats)


if __name__ == '__main__':
    main()
